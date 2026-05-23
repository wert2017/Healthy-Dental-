import os
import time
# Zona horaria Ecuador (Quito) — debe ir antes de cualquier import de datetime
os.environ['TZ'] = 'America/Guayaquil'
try:
    time.tzset()  # Solo funciona en Linux/Mac (Railway usa Linux)
except AttributeError:
    pass  # Windows: no hace falta, el server local usa la hora del sistema

from fastapi import FastAPI, Depends, HTTPException, Query, status, Request, Form, UploadFile, File
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from database import engine, create_db_and_tables, get_session
from models import Paciente, Doctor, Sucursal, Tratamiento, Atencion, AtencionDetalle, Pago, User, TratamientoEnCurso, Insumo, Receta, Proveedor, InventarioSucursal, InventarioDoctor, AuditoriaAtencion, Gasto, HistorialAbono, CategoriaGasto
from sqlmodel import Field, Session, SQLModel, select, create_engine, Relationship
from pydantic import BaseModel
from fastapi.staticfiles import StaticFiles
from sqladmin import Admin, ModelView, BaseView, expose
from sqladmin.authentication import AuthenticationBackend
from starlette.middleware.sessions import SessionMiddleware
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from starlette.responses import RedirectResponse
import uvicorn
from typing import List, Optional
from decimal import Decimal
from datetime import datetime, timedelta, time
from passlib.context import CryptContext
from jose import JWTError, jwt
import secrets
from sqlalchemy import func
from sqlalchemy.orm import selectinload
from wtforms import SelectField
import wtforms

import locale
try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
except:
    pass # Fallback if system doesn't have it

app = FastAPI(title="Clinica HD API")

# --- CORS CONFIGURATION ---
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Fix for HTTPS behind proxy (Railway)
@app.middleware("http")
async def fix_proxy_headers(request: Request, call_next):
    # If the request comes through a proxy with HTTPS, force the scheme to https
    # This ensures url_for and other helpers generate https links, fixing CSS/JS issues
    if request.headers.get("x-forwarded-proto") == "https":
        request.scope["scheme"] = "https"
    response = await call_next(request)
    return response

# --- ROUTERS ---
from inventario_router import router as inventario_router
app.include_router(inventario_router)

# --- AUTH SECURITY ---
SECRET_KEY = "super_secret_key_change_this_in_production"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 # 24 Hours

# Switch to pbkdf2_sha256 to avoid bcrypt/passlib version conflict on Windows
pwd_context = CryptContext(schemes=["pbkdf2_sha256"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(token: str = Depends(oauth2_scheme), session: Session = Depends(get_session)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        sucursal_id_token: Optional[int] = payload.get("sucursal_id") # Read from JWT
        if username is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    user = session.exec(select(User).where(User.username == username)).first()
    if user is None:
        raise credentials_exception
    
    # CRITICAL: Overwrite the DB sucursal_id with the one from the Token
    # This allows admins to have a 'sucursal_id' in memory during their session
    if sucursal_id_token:
        user.sucursal_id = sucursal_id_token
        
    return user

from sqlalchemy import text

@app.on_event("startup")
def on_startup():
    create_db_and_tables()
    
    # Auto-migration: cada columna en su propia transaccion para evitar que un fallo
    # en PostgreSQL deje la transaccion en estado abortado e impida las migraciones siguientes
    migrations = [
        "ALTER TABLE paciente ADD COLUMN sexo VARCHAR;",
        "ALTER TABLE paciente ADD COLUMN edad INTEGER;",
        "ALTER TABLE paciente ADD COLUMN ciudad VARCHAR;",
        "ALTER TABLE historialabono ADD COLUMN atencion_id INTEGER REFERENCES atencion(id);",
    ]
    for sql in migrations:
        try:
            with Session(engine) as session:
                session.execute(text(sql))
                session.commit()
        except Exception:
            pass
            
    # Create Default Admin User (If it doesn't exist)
    with Session(engine) as session:
        try:
            user = session.exec(select(User).where(User.username == "admin")).first()
            if not user:
                print("Creating default admin user...")
                hashed = pwd_context.hash("admin")
                admin_user = User(username="admin", hashed_password=hashed, role="admin")
                session.add(admin_user)
                session.commit()
            else:
                # Do NOT reset password in production/live environments
                pass
        except Exception as e:
            session.rollback()
            print(f"Skipping admin creation error: {e}")
    
    # Re-enable safe auto-seeding of clinics if database is empty
    with Session(engine) as session:
        seed_data(session)


@app.get("/api/public/sucursales")
def list_public_sucursales(session: Session = Depends(get_session)):
    """Publicly list sucursales for the login screen."""
    return session.exec(select(Sucursal)).all()

@app.post("/token")
async def login_for_access_token(
    response: JSONResponse, 
    form_data: OAuth2PasswordRequestForm = Depends(), 
    sucursal_id: Optional[int] = Form(None), # Added for clinic selection
    session: Session = Depends(get_session)
):
    user = session.exec(select(User).where(User.username == form_data.username)).first()
    if not user or not verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    # Validate Sucursal selection if Provided
    final_sucursal_id = sucursal_id
    if user.role != "admin":
        if user.sucursal_id and sucursal_id and user.sucursal_id != sucursal_id:
            raise HTTPException(status_code=403, detail="No tienes acceso a esta sucursal")
        # If user has a fixed sucursal and didn't provide one, use their fixed one
        if user.sucursal_id and not sucursal_id:
            final_sucursal_id = user.sucursal_id
    
    if not final_sucursal_id:
         # For admins or users without a fixed sucursal who didn't pick one, we might need a default or error
         # Let's say we require it if available
         pass

    # Set role cookie for Admin Panel seamless access
    response.set_cookie(key="user_role", value=user.role, httponly=False)
    if final_sucursal_id:
        response.set_cookie(key="sucursal_id", value=str(final_sucursal_id), httponly=False) 
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={
            "sub": user.username,
            "role": user.role,
            "doctor_id": user.doctor_id,
            "sucursal_id": final_sucursal_id
        }, 
        expires_delta=access_token_expires
    )
    
    # Get Doctor Full Name if applicable
    full_name = user.username
    if user.doctor_id:
        doc = session.get(Doctor, user.doctor_id)
        if doc:
            full_name = f"{doc.nombres} {doc.apellidos}"

    sucursal_nombre = None
    if final_sucursal_id:
        suc = session.get(Sucursal, final_sucursal_id)
        if suc:
            sucursal_nombre = suc.nombre

    return {
        "access_token": access_token,
        "token_type": "bearer",
        "role": user.role,
        "doctor_id": user.doctor_id,
        "sucursal_id": final_sucursal_id,
        "sucursal_nombre": sucursal_nombre,
        "username": user.username,
        "full_name": full_name
    }

# Main redirects
@app.get("/")
def root():
    # Serve landing page if built, otherwise redirect to reception
    # Use absolute paths relative to this file to be robust
    current_dir = os.path.dirname(os.path.abspath(__file__))
    landing_index = os.path.join(current_dir, "..", "landing_page", "dist", "index.html")
    
    if os.path.exists(landing_index):
        return FileResponse(landing_index)
    
    print(f"DEBUG: Landing page not found at {landing_index}")
    return RedirectResponse("/recepcion")

# Mount React Landing Page (Assets)
current_dir = os.path.dirname(os.path.abspath(__file__))
LANDING_DIR = os.path.join(current_dir, "..", "landing_page", "dist")
if os.path.exists(LANDING_DIR):
    app.mount("/assets", StaticFiles(directory=os.path.join(LANDING_DIR, "assets")), name="assets")

# Protected Routes - Pages
@app.get("/login")
def login_page():
    return FileResponse("static/login.html")

@app.get("/recepcion", response_class=HTMLResponse)
def recepcion_page(): # Not strictly protecting the HTML serving, but the data inside will be protected. Ideally protect this too.
    # To protect HTML serving, we need to check cookie or token, but standard OAuth2 + Bearer is for API.
    # For simplicity, we serve the HTML publicly, but if you have no token, the API calls inside will fail and redirect you.
    return FileResponse("static/recepcion.html")

@app.get("/recepcion/editar/{atencion_id}", response_class=HTMLResponse)
def editar_atencion_page(atencion_id: int):
    return FileResponse("static/editar.html")

@app.get("/recepcion/imprimir/{atencion_id}", response_class=HTMLResponse)
def imprimir_atencion_page(atencion_id: int):
    return FileResponse("static/imprimir.html") 

@app.get("/gastos", response_class=HTMLResponse)
def gastos_page():
    return FileResponse("static/gastos.html")

@app.get("/cuadre", response_class=HTMLResponse)
def cuadre_page():
    return FileResponse("static/cuadre.html")

@app.get("/admin/inventario", response_class=HTMLResponse)
def admin_inventario_page():
    return FileResponse("static/inventario.html") 

@app.get("/doctor/dashboard", response_class=HTMLResponse)
def doctor_dashboard_page():
     return FileResponse("static/doctor_dashboard.html") 


# --- ADMIN VIEWS ---
class PacienteAdmin(ModelView, model=Paciente):
    name = "Paciente"
    name_plural = "Pacientes"
    icon = "fa-solid fa-user"
    
    # List View configuration
    column_list = [Paciente.historia_clinica, Paciente.numero_identificacion, Paciente.nombres, Paciente.apellidos, Paciente.telefono, Paciente.saldo_favor]
    column_searchable_list = [Paciente.nombres, Paciente.apellidos, Paciente.numero_identificacion]
    
    # Form Configuration (Create/Edit)
    form_columns = [
        Paciente.historia_clinica,
        Paciente.nombres,
        Paciente.apellidos,
        Paciente.edad,
        Paciente.sexo,
        Paciente.tipo_identificacion,
        Paciente.numero_identificacion,
        Paciente.razon_social,
        Paciente.telefono,
        Paciente.email,
        Paciente.saldo_favor
    ]

    # Restrict choices for Dropdowns
    form_overrides = {
        "tipo_identificacion": SelectField,
        "sexo": SelectField
    }
    form_widget_args = {}
    form_args = {
        "tipo_identificacion": {
            "choices": [("CEDULA", "CEDULA"), ("RUC", "RUC"), ("S/N", "S/N")],
            "label": "Tipo Identificacion"
        },
        "sexo": {
            "choices": [("", "-- Seleccionar --"), ("Masculino", "Masculino"), ("Femenino", "Femenino"), ("Otro", "Otro")],
            "label": "Sexo / Género",
            "validators": [wtforms.validators.Optional()]
        },
        "edad": {
            "label": "Edad",
            "validators": [wtforms.validators.Optional()]
        },
        "historia_clinica": {
            "label": "Historia Clínica / Ficha",
            "validators": [wtforms.validators.Optional()]
        },
        "apellidos": {
            "validators": [wtforms.validators.Optional()]
        },
        "numero_identificacion": {
            "validators": [wtforms.validators.Optional()]
        },
        "telefono": {
            "validators": [wtforms.validators.Optional()]
        }
    }

    def list_query(self, request: Request):
        stmt = super().list_query(request)
        sucursal_id = int(request.cookies.get("sucursal_id", "1"))
        stmt = stmt.where(Paciente.sucursal_id == sucursal_id)
        return stmt

    def count_query(self, request: Request):
        stmt = super().count_query(request)
        sucursal_id = int(request.cookies.get("sucursal_id", "1"))
        stmt = stmt.where(Paciente.sucursal_id == sucursal_id)
        return stmt

    # We explicitly exclude 'atenciones' and 'fecha_creacion' from the form

    async def on_model_change(self, data, model, is_created, request: Request):
        from sqlmodel import Session, select
        import uuid
        
        # 1. Ensure the patient belongs to the current admin's branch
        admin_sucursal_id = int(request.cookies.get("sucursal_id", "1"))
        if is_created and not model.sucursal_id:
            model.sucursal_id = admin_sucursal_id
        
        # Default missing values to bypass DB constraints safely
        if not data.get('tipo_identificacion'):
            data['tipo_identificacion'] = "S/N"
            model.tipo_identificacion = "S/N"
        if not data.get('numero_identificacion'):
            generated = f"SD-{str(uuid.uuid4())[:8]}"
            data['numero_identificacion'] = generated
            model.numero_identificacion = generated
        if not data.get('apellidos'):
            data['apellidos'] = ""
            model.apellidos = ""
        if not data.get('telefono'):
            data['telefono'] = ""
            model.telefono = ""
        
        # 2. Auto-generate Clinical History Number if missing (supports manual override)
        if not model.historia_clinica or str(model.historia_clinica).strip() == "":
            from database import engine as db_engine
            with Session(db_engine) as session:
                # Find the Sucursal to get prefix
                from models import Sucursal
                suc = session.get(Sucursal, model.sucursal_id)
                prefix = suc.nombre[:3].upper() if suc and len(suc.nombre) >= 3 else "GEN"
                
                # Find the highest HC number already used for THIS branch specifically
                # to avoid jumping to a global ID that belongs to other branches.
                existing_hcs = session.exec(
                    select(Paciente.historia_clinica)
                    .where(Paciente.sucursal_id == model.sucursal_id)
                    .where(Paciente.historia_clinica.ilike(f"HC-{prefix}-%"))
                ).all()
                max_num = 0
                for hc in existing_hcs:
                    try:
                        num = int(hc.split("-")[-1])
                        if num > max_num:
                            max_num = num
                    except (ValueError, IndexError):
                        pass
                new_number = max_num + 1

                model.historia_clinica = f"HC-{prefix}-{new_number:04d}"

class DoctorAdmin(ModelView, model=Doctor):
    column_list = [Doctor.apellidos, Doctor.nombres, Doctor.cedula, Doctor.sucursal, Doctor.activo]
    form_columns = [Doctor.nombres, Doctor.apellidos, Doctor.cedula, Doctor.telefono, Doctor.email, Doctor.sucursal, Doctor.activo]

    def is_accessible(self, request: Request) -> bool:
        return request.cookies.get("user_role") == "admin"
    
    def is_visible(self, request: Request) -> bool:
        return request.cookies.get("user_role") == "admin"

class SucursalAdmin(ModelView, model=Sucursal):
    column_list = [Sucursal.nombre, Sucursal.direccion]
    form_columns = [Sucursal.nombre, Sucursal.direccion]
    icon = "fa-solid fa-building"

    def is_accessible(self, request: Request) -> bool:
        return request.cookies.get("user_role") == "admin"
    
    def is_visible(self, request: Request) -> bool:
        return request.cookies.get("user_role") == "admin"

class UserAdmin(ModelView, model=User):
    name = "Usuario"
    name_plural = "Usuarios"
    column_list = [User.username, User.role, User.doctor, User.sucursal]
    form_columns = [User.username, "hashed_password", User.role, User.doctor, User.sucursal]
    icon = "fa-solid fa-users-gear"
    
    form_args = {
        "hashed_password": {"label": "Contraseña / Password"}
    }

    def is_accessible(self, request: Request) -> bool:
        return request.cookies.get("user_role") == "admin"
    
    def is_visible(self, request: Request) -> bool:
        return request.cookies.get("user_role") == "admin"

    async def on_model_change(self, data, model, is_created, request: Request):
        # Si se envió una contraseña (en creación o edición), la hasheamos
        if "hashed_password" in data and data["hashed_password"]:
            # Solo hasheamos si no parece ser ya un hash (para evitar doble hash en ediciones sin cambio)
            # Aunque lo ideal es que el formulario no muestre el hash.
            # Por ahora, si el usuario escribe algo, lo hasheamos.
            data["hashed_password"] = pwd_context.hash(data["hashed_password"])

class TratamientoAdmin(ModelView, model=Tratamiento):
    column_list = [Tratamiento.codigo, Tratamiento.nombre, Tratamiento.precio_base, Tratamiento.activo]
    form_columns = [Tratamiento.codigo, Tratamiento.nombre, Tratamiento.precio_base, Tratamiento.activo]

class AtencionAdmin(ModelView, model=Atencion):
    name = "Atención"
    name_plural = "Atenciones"
    icon = "fa-solid fa-notes-medical"
    column_list = [Atencion.id, Atencion.fecha, Atencion.paciente, Atencion.doctor, Atencion.estado, Atencion.validado]
    column_details_list = [Atencion.id, Atencion.fecha, Atencion.paciente, Atencion.doctor, Atencion.estado, Atencion.validado, Atencion.observaciones, Atencion.detalles, Atencion.pagos]
    can_create = True
    form_columns = [Atencion.fecha, Atencion.paciente, Atencion.doctor, Atencion.observaciones]

    def is_accessible(self, request) -> bool:
        return request.cookies.get("user_role") == "admin"

    def is_visible(self, request: Request) -> bool:
        return request.cookies.get("user_role") == "admin"

    async def on_model_change(self, data, model, is_created, request: Request):
        if is_created:
            sucursal_id = request.cookies.get("sucursal_id")
            if sucursal_id:
                data["sucursal_id"] = int(sucursal_id)

class AtencionDetalleAdmin(ModelView, model=AtencionDetalle):
    name = "Detalle de Atención"
    name_plural = "Detalles de Atenciones"
    icon = "fa-solid fa-list-check"
    column_list = [AtencionDetalle.id, AtencionDetalle.atencion_id, AtencionDetalle.tratamiento, AtencionDetalle.doctor, AtencionDetalle.cantidad, AtencionDetalle.precio_unitario]
    form_columns = [AtencionDetalle.atencion, AtencionDetalle.tratamiento, AtencionDetalle.doctor, AtencionDetalle.cantidad, AtencionDetalle.precio_unitario, AtencionDetalle.porcentaje_comision]
    can_create = True
    can_edit = True
    can_delete = True

    def is_accessible(self, request: Request) -> bool:
        return request.cookies.get("user_role") == "admin"

    def is_visible(self, request: Request) -> bool:
        return request.cookies.get("user_role") == "admin"

class PagoAdmin(ModelView, model=Pago):
    column_list = [Pago.id, Pago.forma_pago, Pago.monto, Pago.atencion_id]

# --- ADMIN HELPERS ---
from markupsafe import Markup

class InsumoAdmin(ModelView, model=Insumo):
    column_list = [Insumo.nombre, Insumo.stock_actual, Insumo.unidad_medida, Insumo.proveedor]
    form_columns = [Insumo.nombre, Insumo.unidad_medida, Insumo.stock_actual, Insumo.stock_minimo, Insumo.proveedor]
    icon = "fa-solid fa-boxes-stacked"

    # Visual Alert for Low Stock (Temporarily Disabled)
    # def stock_formatter(view, context, model, name):
    #     try:
    #         val = model.stock_actual if model.stock_actual is not None else 0
    #         min_val = model.stock_minimo if model.stock_minimo is not None else 0
    #         
    #         if val <= min_val:
    #             return Markup(f'<span style="color: red; font-weight: bold;">{val} (Bajo)</span>')
    #         return str(val)
    #     except Exception:
    #         return str(model.stock_actual)

    # column_formatters = {
    #     "stock_actual": stock_formatter
    # }

class RecetaAdmin(ModelView, model=Receta):
    name = "Receta (Configuración)"
    name_plural = "Configuración de Recetas"
    column_list = [Receta.tratamiento, Receta.insumo, Receta.cantidad_requerida]
    icon = "fa-solid fa-prescription-bottle-medical"

class ProveedorAdmin(ModelView, model=Proveedor):
    column_list = [Proveedor.nombre, Proveedor.contacto, Proveedor.telefono, Proveedor.email, Proveedor.activo]
    form_columns = [Proveedor.nombre, Proveedor.contacto, Proveedor.telefono, Proveedor.email, Proveedor.activo]
    icon = "fa-solid fa-truck"

class InventarioSucursalAdmin(ModelView, model=InventarioSucursal):
    name = "Stock por Sucursal"
    name_plural = "Stock Sucursales"
    column_list = [InventarioSucursal.sucursal, InventarioSucursal.insumo, InventarioSucursal.stock_actual]
    form_columns = [InventarioSucursal.sucursal, InventarioSucursal.insumo, InventarioSucursal.stock_actual, InventarioSucursal.stock_minimo]
    icon = "fa-solid fa-store"

class InventarioDoctorAdmin(ModelView, model=InventarioDoctor):
    name = "Stock Personal (Doctor)"
    name_plural = "Stock Personal Doctores"
    column_list = ["doctor", "insumo", "stock_actual"]
    icon = "fa-solid fa-user-doctor"

class CategoriaGastoAdmin(ModelView, model=CategoriaGasto):
    name = "Categoría de Gasto"
    name_plural = "Categorías de Gasto"
    column_list = [CategoriaGasto.nombre, CategoriaGasto.activo]
    form_columns = [CategoriaGasto.nombre, CategoriaGasto.activo]
    icon = "fa-solid fa-tags"

    def is_accessible(self, request):
        return request.cookies.get("user_role") == "admin"

class MovimientosLink(BaseView):
    name = "Movimientos de Inventario"
    icon = "fa-solid fa-truck-ramp-box"
    
    @expose("/movimientos", methods=["GET"])
    def movements_page(self, request):
        return RedirectResponse(url="/static/inventario.html")

# --- APP STARTUP & SEEDING ---
def seed_data(session: Session):
    # Seed Sucursales (Clinics)
    target_sucursales = [
        {"nombre": "HEALTHY DENTAL LA MAGDALENA", "direccion": "Sur"},
        {"nombre": "Sucursal Norte", "direccion": "Calle Norte 456"},
    ]
    for s_data in target_sucursales:
        try:
            exists = session.exec(select(Sucursal).where(Sucursal.nombre == s_data["nombre"])).first()
            if not exists:
                session.add(Sucursal(**s_data))
                session.commit()
        except Exception:
            session.rollback()

    # Seed Treatments
    if not session.exec(select(Tratamiento)).first():
        tratamientos = [
            Tratamiento(codigo="LDC", nombre="Lente de contacto", precio_base=150.00),
            Tratamiento(codigo="CAR", nombre="Carilla", precio_base=80.00),
            Tratamiento(codigo="BLA", nombre="Blanqueamiento", precio_base=120.00),
            Tratamiento(codigo="LIM", nombre="Limpieza", precio_base=30.00),
            Tratamiento(codigo="CAL", nombre="Calza", precio_base=40.00),
            Tratamiento(codigo="COR", nombre="Corona", precio_base=200.00),
            Tratamiento(codigo="EXT", nombre="Extraccion", precio_base=25.00),
            Tratamiento(codigo="END", nombre="Endodoncia", precio_base=180.00),
            Tratamiento(codigo="PRO", nombre="Protesis", precio_base=350.00),
        ]
        for t in tratamientos:
            session.add(t)
        print("SEED: Tratamientos agregados.")

    # Seed Doctors
    if not session.exec(select(Doctor)).first():
        doctores = [
            Doctor(nombres="Juan", apellidos="Perez", cedula="0900000001", telefono="0991234567"),
            Doctor(nombres="Ana", apellidos="Gomez", cedula="0900000002", telefono="0997654321"),
        ]
        for d in doctores:
            session.add(d)
        print("SEED: Doctores agregados.")

    # Seed Patients
    if not session.exec(select(Paciente)).first():
        pacientes = [
            Paciente(nombres="Carlos", apellidos="Ruiz", numero_identificacion="0999999999", tipo_identificacion="CEDULA", telefono="0987654321", email="carlos@example.com", historia_clinica="HC-1001"),
            Paciente(nombres="Maria", apellidos="Lopez", numero_identificacion="0988888888", tipo_identificacion="CEDULA", telefono="0911223344", email="maria@example.com", historia_clinica="HC-1002"),
            Paciente(nombres="Jose", apellidos="Vera", numero_identificacion="0977777777", tipo_identificacion="CEDULA", telefono="0955667788", email="jose@example.com", historia_clinica="HC-1003"),
        ]
        for p in pacientes:
            session.add(p)
        print("SEED: Pacientes agregados.")
    
    # Seed expense categories
    if not session.exec(select(CategoriaGasto)).first():
        categorias_default = [
            CategoriaGasto(nombre="GENERAL"),
            CategoriaGasto(nombre="NÓMINA"),
            CategoriaGasto(nombre="SUELDO FIJO"),
            CategoriaGasto(nombre="COMISIONES"),
            CategoriaGasto(nombre="INSUMOS"),
            CategoriaGasto(nombre="MANTENIMIENTO"),
            CategoriaGasto(nombre="RETIRO SOCIOS"),
            CategoriaGasto(nombre="SERVICIOS BÁSICOS"),
            CategoriaGasto(nombre="OTROS"),
        ]
        for c in categorias_default:
            session.add(c)
        session.commit()
        print("SEED: Categorías de gasto agregadas.")
    else:
        # Ensure SUELDO FIJO and COMISIONES exist even on existing installs
        for nombre_nuevo in ["SUELDO FIJO", "COMISIONES"]:
            if not session.exec(select(CategoriaGasto).where(CategoriaGasto.nombre == nombre_nuevo)).first():
                session.add(CategoriaGasto(nombre=nombre_nuevo))
        session.commit()

    session.commit()

# Startup event merged into the primary on_startup handler above
# @app.on_event("startup")
# def on_startup_redundant():
#     create_db_and_tables()
#     with Session(engine) as session:
#         seed_data(session)

# --- ADMIN SETUP ---
current_dir = os.path.dirname(os.path.abspath(__file__))
templates_path = os.path.join(current_dir, "templates")

admin = Admin(app, engine, title="Admin - Clínica HD", templates_dir=templates_path)
# admin.add_view(BackToReception) # Custom view removed in favor of template override
admin.add_view(PacienteAdmin)
admin.add_view(DoctorAdmin)
admin.add_view(TratamientoAdmin)
admin.add_view(AtencionAdmin)
admin.add_view(AtencionDetalleAdmin)
admin.add_view(PagoAdmin)
admin.add_view(SucursalAdmin) 
admin.add_view(InsumoAdmin)
admin.add_view(ProveedorAdmin)
admin.add_view(InventarioSucursalAdmin)
admin.add_view(InventarioDoctorAdmin)
admin.add_view(UserAdmin)
admin.add_view(MovimientosLink)
admin.add_view(RecetaAdmin)
admin.add_view(CategoriaGastoAdmin)

# --- CATEGORÍAS DE GASTO ---
@app.get("/api/categorias-gasto")
def list_categorias_gasto(session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    cats = session.exec(
        select(CategoriaGasto).where(CategoriaGasto.activo == True).order_by(CategoriaGasto.nombre)
    ).all()
    return [{"id": c.id, "nombre": c.nombre} for c in cats]

# --- USER LIST FOR DROPDOWNS ---
@app.get("/api/usuarios")
def list_usuarios(session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    """Returns a list of active users to be used in dropdowns (e.g. Responsable de Gasto)"""
    usuarios = session.exec(select(User).order_by(User.username.asc())).all()
    # Return minimal data for security
    return [{"id": u.id, "username": u.username} for u in usuarios]

# --- STATIC FILES ---
app.mount("/static", StaticFiles(directory="static"), name="static")


# --- HELPER FUNCTIONS ---
def process_stock_deduction(atencion: Atencion, session: Session, sucursal_id: int):
    """
    Deducts stock based on treatments in the attention.
    Should be called ONLY when validating an attention.
    Now supports Multi-Clinic logic: Deducts from InventarioSucursal.
    """
    if not sucursal_id:
        print("WARNING: No sucursal_id provided for stock deduction. Skipping.")
        return

    deductions_made = []
    
    for detalle in atencion.detalles:
        if detalle.tratamiento and detalle.tratamiento.insumos_requeridos:
            # Use the doctor from the detail (or the main doctor of the attention)
            current_doctor_id = detalle.doctor_id if detalle.doctor_id else atencion.doctor_id
            
            for receta in detalle.tratamiento.insumos_requeridos:
                cantidad_a_descontar = receta.cantidad_requerida * detalle.cantidad
                
                # 1. ATTEMPT PRIORITY: Personal Doctor stock (Level 3)
                inv_doctor = None
                if current_doctor_id:
                    inv_doctor = session.exec(select(InventarioDoctor).where(
                        InventarioDoctor.doctor_id == current_doctor_id,
                        InventarioDoctor.insumo_id == receta.insumo_id
                    )).first()

                if inv_doctor and inv_doctor.stock_actual >= cantidad_a_descontar:
                    # Deduct from Doctor personal stock
                    inv_doctor.stock_actual -= cantidad_a_descontar
                    session.add(inv_doctor)
                    deductions_made.append(f"{receta.insumo.nombre} (-{cantidad_a_descontar}) @ Doctor {current_doctor_id}")
                else:
                    # 2. FALLBACK: Sucursal Inventory (Level 2)
                    inv_suc = session.exec(select(InventarioSucursal).where(
                        InventarioSucursal.sucursal_id == sucursal_id,
                        InventarioSucursal.insumo_id == receta.insumo_id
                    )).first()
                    
                    if inv_suc:
                        inv_suc.stock_actual -= cantidad_a_descontar
                        session.add(inv_suc)
                        deductions_made.append(f"{receta.insumo.nombre} (-{cantidad_a_descontar}) @ Sucursal {sucursal_id}")
                    else:
                        # Fallback Create negative record if not exists in sucursal
                        new_inv = InventarioSucursal(
                            sucursal_id=sucursal_id,
                            insumo_id=receta.insumo_id,
                            stock_actual=-cantidad_a_descontar,
                            stock_minimo=0
                        )
                        session.add(new_inv)
                        deductions_made.append(f"{receta.insumo.nombre} (CREATED -{cantidad_a_descontar}) @ Sucursal {sucursal_id}")
    
    if deductions_made:
        print(f"STOCK UPDATE: Atencion {atencion.id} consumed: {', '.join(deductions_made)}")

# --- API ROUTES ---

@app.get("/api/admin/proxima-ficha")
def get_proxima_ficha(request: Request, session: Session = Depends(get_session)):
    sucursal_id_cookie = request.cookies.get("sucursal_id")
    sucursal_id = int(sucursal_id_cookie) if sucursal_id_cookie else 1
    
    suc = session.get(Sucursal, sucursal_id)
    prefix = suc.nombre[:3].upper() if suc and len(suc.nombre) >= 3 else "GEN"
    
    existing_hcs = session.exec(
        select(Paciente.historia_clinica)
        .where(Paciente.sucursal_id == sucursal_id)
        .where(Paciente.historia_clinica.ilike(f"HC-{prefix}-%"))
    ).all()
    max_num = 0
    for hc in existing_hcs:
        try:
            num = int(hc.split("-")[-1])
            if num > max_num:
                max_num = num
        except (ValueError, IndexError):
            pass
    new_number = max_num + 1
    
    return {"proxima_ficha": f"HC-{prefix}-{new_number:04d}"}

@app.get("/api/pacientes/buscar-doctor")
def buscar_pacientes_doctor(q: str = "", session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    """Búsqueda de pacientes para el doctor: devuelve solo HC, género y edad."""
    if not q or len(q.strip()) < 2:
        return []
    terms = q.strip().split()
    statement = select(Paciente).where(Paciente.sucursal_id == user.sucursal_id)
    for term in terms:
        t = f"%{term}%"
        statement = statement.where(
            (Paciente.nombres.ilike(t)) | (Paciente.apellidos.ilike(t))
        )
    pacientes = session.exec(statement.limit(20)).all()
    return [
        {
            "historia_clinica": p.historia_clinica,
            "nombres": p.nombres,
            "apellidos": p.apellidos,
            "sexo": p.sexo or "—",
            "edad": p.edad,
        }
        for p in pacientes
    ]


@app.get("/api/pacientes")
def search_pacientes(q: str = "", session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    if not q:
        return []
    
    terms = q.strip().split()
    statement = select(Paciente).where(Paciente.sucursal_id == user.sucursal_id)
    
    for term in terms:
        t = f"%{term}%"
        statement = statement.where(
            (Paciente.nombres.ilike(t)) |
            (Paciente.apellidos.ilike(t)) |
            (Paciente.numero_identificacion.ilike(t)) |
            (Paciente.historia_clinica.ilike(t))
        )
        
    statement = statement.limit(10)
    pacientes = session.exec(statement).all()
    
    # Calculate financial status for each patient
    results_data = []
    for p in pacientes:
        total_consumido = sum([sum([d.total_calculado for d in a.detalles]) for a in p.atenciones])
        total_pagado = sum([sum([pay.monto for pay in a.pagos]) for a in p.atenciones])
        balance = total_pagado - total_consumido
        
        status = "AL_DIA"
        if balance < -0.01:
            status = "DEUDA"
        elif balance > 0.01:
            status = "FAVOR"
            
        results_data.append({
            "id": p.id,
            "nombres": p.nombres,
            "apellidos": p.apellidos,
            "numero_identificacion": p.numero_identificacion,
            "historia_clinica": p.historia_clinica,
            "telefono": p.telefono,
            "email": p.email,
            "fecha_creacion": p.fecha_creacion,
            "total_pagado": total_pagado,
            "balance": balance,
            "estado_financiero": status,
            "saldo_favor": p.saldo_favor
        })
        
    return results_data








@app.delete("/api/admin/nuke-pacientes")
def nuke_all_pacientes(session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    if user.role != "admin" or user.username != "admin":
        raise HTTPException(status_code=403, detail="Unicamente el Super-Administrador puede ejecutar esta accion")
    from sqlalchemy import text
    try:
        session.exec(text("DELETE FROM auditoriaatencion"))
        session.exec(text("DELETE FROM historialabono"))
        session.exec(text("DELETE FROM pago"))
        session.exec(text("DELETE FROM atenciondetalle"))
        session.exec(text("DELETE FROM atencion"))
        session.exec(text("DELETE FROM tratamientoencurso"))
        session.exec(text("DELETE FROM gasto"))
        session.exec(text("DELETE FROM paciente"))
        session.commit()
        return {"status": "success", "message": "Base de datos limpia. Se conservaron sucursales, doctores, tratamientos y usuarios."}
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/pacientes/validar-excel")
async def validar_excel_pacientes(file: UploadFile = File(...), user: User = Depends(get_current_user)):
    import openpyxl, io

    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="El archivo debe ser un documento Excel (.xlsx)")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    sheet = wb.active

    headers = {str(cell.value).strip(): idx for idx, cell in enumerate(sheet[1]) if cell.value is not None}

    def clean_str(val):
        if val is None:
            return None
        s = str(val).strip()
        return s if s else None

    def get_col(row, col_name):
        if col_name in headers:
            idx = headers[col_name]
            return row[idx] if idx < len(row) else None
        return None

    total = 0
    validas = 0
    sin_nombre = []      # filas sin apellido NI nombre
    nombre_incompleto = []  # filas con solo uno de los dos

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Skip completely empty rows
        if all(c is None for c in row):
            continue
        total += 1

        apellidos = clean_str(get_col(row, 'APELLIDOS'))
        nombres   = clean_str(get_col(row, 'NOMBRE'))
        cedula    = clean_str(get_col(row, 'CEDULA'))
        ficha     = clean_str(get_col(row, 'FICHA'))

        entry = {
            "fila": row_idx,
            "ficha": ficha or "-",
            "cedula": cedula or "-",
            "apellidos": apellidos or "",
            "nombres": nombres or "",
        }

        if not apellidos and not nombres:
            sin_nombre.append(entry)
        elif not apellidos or not nombres:
            nombre_incompleto.append(entry)
            validas += 1
        else:
            validas += 1

    return {
        "total": total,
        "validas": validas,
        "sin_nombre": len(sin_nombre),
        "nombre_incompleto": len(nombre_incompleto),
        "muestra_sin_nombre": sin_nombre[:20],       # primeras 20 para mostrar en UI
        "muestra_incompleto": nombre_incompleto[:20],
        "columnas_detectadas": list(headers.keys()),
    }


@app.post("/api/pacientes/importar-excel")
async def importar_pacientes_excel(file: UploadFile = File(...), sucursal_id: Optional[int] = Form(None), session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    import openpyxl
    import io
    
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="El archivo debe ser un documento Excel (.xlsx)")
        
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        sheet = wb.active
        
        # Read the headers from the first row
        headers = {str(cell.value).strip(): idx for idx, cell in enumerate(sheet[1]) if cell.value is not None}
        
        target_sucursal_id = sucursal_id or user.sucursal_id
        suc = session.get(Sucursal, target_sucursal_id)
        prefix = suc.nombre[:3].upper() if suc and len(suc.nombre) >= 3 else "GEN"
        last_patient = session.exec(select(Paciente).order_by(Paciente.id.desc())).first()
        hc_counter = (last_patient.id + 1) if last_patient else 1
        prov_counter = 1
        inserted = 0
        skipped = 0
        
        def clean_str(val):
            if val is None:
                return None
            s = str(val).strip()
            return s if s else None
            
        # Iterate over the rows, skipping the header (row 2 onwards)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            def get_col(col_name):
                if col_name in headers:
                    idx = headers[col_name]
                    if idx < len(row):
                        return row[idx]
                return None

            apellidos = clean_str(get_col('APELLIDOS'))
            nombres = clean_str(get_col('NOMBRE'))
            
            if not apellidos and not nombres:
                skipped += 1
                continue
                
            apellidos = apellidos or ""
            nombres = nombres or ""
            sexo = clean_str(get_col('SEXO'))
            
            edad_val = get_col('EDAD')
            edad = None
            if edad_val is not None:
                try:
                    edad = int(float(edad_val))
                except (ValueError, TypeError):
                    pass
                    
            telefono = clean_str(get_col('TELEFONO'))
            if not telefono:
                telefono = "0999999999"
                
            cedula = clean_str(get_col('CEDULA'))
            if not cedula or len(cedula) < 4:
                cedula = f"PROV-{prov_counter:05d}"
                prov_counter += 1
                
            # Verificar si existe cédula en base de datos
            existing = session.exec(select(Paciente).where(Paciente.numero_identificacion == cedula)).first()
            if existing:
                skipped += 1
                continue
                
            ciudad = clean_str(get_col('CIUDAD'))
            
            ficha_val = clean_str(get_col('FICHA'))
            if ficha_val:
                try:
                    num_ficha = int(float(ficha_val))
                    historia_clinica = f"HC-{prefix}-{num_ficha:04d}"
                except (ValueError, TypeError):
                    # Si tiene letras u otros caracteres raros
                    historia_clinica = f"HC-{prefix}-{ficha_val}"
            else:
                historia_clinica = f"HC-{prefix}-{hc_counter:04d}"
                hc_counter += 1
            
            nuevo_paciente = Paciente(
                tipo_identificacion="CED" if not cedula.startswith("PROV") else "PROV",
                numero_identificacion=cedula,
                nombres=nombres,
                apellidos=apellidos,
                historia_clinica=historia_clinica,
                telefono=telefono,
                email=None,
                sexo=sexo,
                edad=edad,
                ciudad=ciudad,
                activo=True,
                sucursal_id=target_sucursal_id
            )
            session.add(nuevo_paciente)
            inserted += 1
            
        session.commit()
        return {"status": "success", "message": f"¡Éxito! {inserted} pacientes importados masivamente.", "inserted": inserted, "skipped": skipped}
        
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=f"Error procesando archivo: {str(e)}")
@app.get("/api/doctores")
def list_doctores(session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    """List active doctors, filtered by the user's sucursal if they are not an admin."""
    query = select(Doctor).where(Doctor.activo == True)
    if user.role != "admin" and user.sucursal_id:
        # Include doctors assigned to the user's sucursal AND doctors with no sucursal (global)
        query = query.where(
            (Doctor.sucursal_id == user.sucursal_id) | (Doctor.sucursal_id == None)
        )
    return session.exec(query).all()

@app.get("/api/vendedores")
def list_vendedores(session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    """Returns non-doctor staff users for the vendedor dropdown (recepcion, admin)."""
    query = select(User).where(User.role != "doctor")
    if user.role != "admin" and user.sucursal_id:
        query = query.where(User.sucursal_id == user.sucursal_id)
    usuarios = session.exec(query).all()
    return [{"id": u.id, "username": u.username} for u in usuarios]

@app.get("/api/tratamientos")
def list_tratamientos(session: Session = Depends(get_session)):
    return session.exec(select(Tratamiento).where(Tratamiento.activo == True)).all()

@app.post("/api/atenciones")
def create_atencion(paciente_id: int, fecha: Optional[str] = None, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    doc_id = user.doctor_id if user.role == "doctor" else None

    fecha_atencion = datetime.now()
    if fecha and user.role == "admin":
        try:
            fecha_atencion = datetime.fromisoformat(fecha)
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD o YYYY-MM-DDTHH:MM")

    atencion = Atencion(
        paciente_id=paciente_id,
        sucursal_id=user.sucursal_id,
        doctor_id=doc_id,
        fecha=fecha_atencion
    )
    session.add(atencion)
    session.commit()
    session.refresh(atencion)
    return atencion

@app.post("/api/atenciones/{atencion_id}/terminar")
def terminar_atencion_clinica(atencion_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    """
    Doctor finishes clinical part. Status -> REALIZADO.
    Does NOT deduct stock yet (Reconciliation does that).
    Does NOT process payments.
    """
    atencion = session.get(Atencion, atencion_id)
    if not atencion:
        raise HTTPException(status_code=404, detail="Atención no encontrada")
    
    if atencion.validado:
        raise HTTPException(status_code=400, detail="Atención ya validada")

    atencion.estado = "REALIZADO"
    session.add(atencion)
    session.commit()
    return {"status": "ok", "estado": "REALIZADO"}

@app.get("/api/doctores")
def get_doctores(session: Session = Depends(get_session)):
    return session.exec(select(Doctor).where(Doctor.activo == True)).all()

# ... (omitted get_dashboard_atenciones, get_atencion_detail, add_detalle, delete_detalle, update_detalle, sync_pagos, delete_atencion) ...

@app.post("/api/atenciones/{atencion_id}/validar")
def validar_atencion(atencion_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    # Eager load relationships for inventory and payment logic
    atencion = session.exec(
        select(Atencion)
        .where(Atencion.id == atencion_id)
        .options(
            selectinload(Atencion.detalles).selectinload(AtencionDetalle.tratamiento).selectinload(Tratamiento.insumos_requeridos).selectinload(Receta.insumo),
            selectinload(Atencion.paciente),
            selectinload(Atencion.pagos) # REQUIRED for Wallet deduction
        )
    ).first()
    
    if atencion:
        if atencion.validado:
             return {"message": "Ya estaba validado"}
             
        atencion.validado = True
        atencion.estado = "FINALIZADO"
        
        # --- INVENTORY DEDUCTION (Multi-Clinic) ---
        # Prefer atencion.sucursal_id (where it was created), fallback to validator's sucursal
        sucursal_target = atencion.sucursal_id if atencion.sucursal_id else user.sucursal_id
        process_stock_deduction(atencion, session, sucursal_target)

        session.add(atencion)
        session.commit()
    return {"message": "Validado con éxito"}

@app.post("/api/atenciones/{atencion_id}/solicitar-revision")
def solicitar_revision_atencion(atencion_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    """
    Doctor marks attention as needing review from receptionist.
    Status -> POR_REVISAR.
    """
    atencion = session.get(Atencion, atencion_id)
    if not atencion:
        raise HTTPException(status_code=404, detail="Atención no encontrada")
    
    if atencion.validado:
        raise HTTPException(status_code=400, detail="Atención ya validada")

    atencion.estado = "POR_REVISAR"
    
    # Audit trail
    log = AuditoriaAtencion(
        atencion_id=atencion.id,
        usuario_id=user.id,
        accion="SOLICITAR_REVISION",
        descripcion="El doctor solicitó revisión a recepción"
    )
    session.add(log)
    
    session.add(atencion)
    session.commit()
    return {"message": "Solicitud de revisión enviada a Recepción", "estado": "POR_REVISAR"}

@app.post("/api/atenciones/{atencion_id}/terminar")
def terminar_atencion_clinica(atencion_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    """
    Doctor finishes clinical part. Status -> REALIZADO.
    Does NOT deduct stock yet (Reconciliation does that).
    Does NOT process payments.
    """
    atencion = session.get(Atencion, atencion_id)
    if not atencion:
        raise HTTPException(status_code=404, detail="Atención no encontrada")
    
    if atencion.validado:
        raise HTTPException(status_code=400, detail="Atención ya validada")

    atencion.estado = "REALIZADO"
    session.add(atencion)
    session.commit()
    return {"status": "ok", "estado": "REALIZADO"}

@app.get("/api/doctores")
def get_doctores(session: Session = Depends(get_session)):
    return session.exec(select(Doctor).where(Doctor.activo == True)).all()

@app.get("/api/atenciones/global")
def get_global_atenciones(
    search: str = None,
    status: str = "all",
    start_date: str = None,
    end_date: str = None,
    page: int = 1,
    size: int = 50,
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user)
):
    """
    Returns all attentions with search, payment status, and date range filtering.
    For Pagos Management (Secretary).
    """
    try:
        skip = (page - 1) * size
        query = select(Atencion).order_by(Atencion.fecha.desc()).options(
            selectinload(Atencion.paciente),
            selectinload(Atencion.detalles).selectinload(AtencionDetalle.tratamiento),
            selectinload(Atencion.pagos)
        )

        # Apply search filter at DB level so pagination doesn't hide results
        if search:
            s = f"%{search}%"
            query = query.join(Paciente, Atencion.paciente_id == Paciente.id).where(
                Paciente.nombres.ilike(s) |
                Paciente.apellidos.ilike(s) |
                Paciente.numero_identificacion.ilike(s)
            )

        # Apply Date Filters at DB level for efficiency
        if start_date:
            try:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                query = query.where(Atencion.fecha >= start_dt)
            except ValueError:
                pass
        if end_date:
            try:
                end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
                query = query.where(Atencion.fecha < end_dt)
            except ValueError:
                pass

        atenciones = session.exec(query.offset(skip).limit(size)).all()

        results = []
        for a in atenciones:
            total_atencion = sum([d.total_calculado for d in a.detalles])
            total_pagado = sum([p.monto for p in a.pagos])
            saldo_pendiente = total_atencion - total_pagado

            # Status Filter
            if status == "pending" and saldo_pendiente <= 0.01:
                continue
            if status == "paid" and saldo_pendiente > 0.01:
                continue

            results.append({
                "id": a.id,
                "fecha": a.fecha,
                "paciente_nombre": f"{a.paciente.nombres} {a.paciente.apellidos}",
                "paciente_id": a.paciente.numero_identificacion,
                "total_atencion": total_atencion,
                "total_pagado": total_pagado,
                "saldo_pendiente": saldo_pendiente,
                "validado": a.validado,
                "estado": a.estado
            })

        return results
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/atenciones/dashboard")
def get_dashboard_atenciones(fecha: str = None, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    try:
        today = datetime.now().date()
        today_start = datetime.combine(today, datetime.min.time())

        target_date = datetime.strptime(fecha, "%Y-%m-%d").date() if fecha else today
        target_start = datetime.combine(target_date, datetime.min.time())
        target_end = datetime.combine(target_date, time(23, 59, 59))

        # Auto-validation only when loading today (past days are already conciliated)
        if target_date == today:
            past_unvalidated = session.exec(select(Atencion).where(Atencion.fecha < today_start).where(Atencion.validado == False)).all()
            for att in past_unvalidated:
                att.validado = True
                att.estado = "FINALIZADO"
                target_sucursal = att.sucursal_id if att.sucursal_id else user.sucursal_id
                process_stock_deduction(att, session, target_sucursal)
                session.add(att)
            if past_unvalidated:
                session.commit()

        # Filter Dashboard: Only return attentions from the selected date AND current branch
        query = (
            select(Atencion)
            .where(Atencion.fecha >= target_start)
            .where(Atencion.fecha <= target_end)
            .where(Atencion.sucursal_id == user.sucursal_id) # REQUIRED for Multi-Clinic separation
            .order_by(Atencion.fecha.desc())
            .options(
                selectinload(Atencion.detalles).selectinload(AtencionDetalle.tratamiento),
                selectinload(Atencion.detalles).selectinload(AtencionDetalle.doctor),
                selectinload(Atencion.paciente),
                selectinload(Atencion.pagos)
            )
        )
        
        if user.role == "doctor" and user.doctor_id:
            # Filter by MAIN doctor OR if they have any treatment in the details
            query = query.join(Atencion.detalles, isouter=True).where(
                (Atencion.doctor_id == user.doctor_id) | 
                (AtencionDetalle.doctor_id == user.doctor_id)
            ).distinct()
            
        atenciones = session.exec(query).all()
        
        results = []
        for a in atenciones:
            # Aggregate treatments and doctors for the summary row
            tratamientos_names = [d.tratamiento.nombre for d in a.detalles if d.tratamiento]
            doctors_names = list(set([d.doctor.nombres + " " + d.doctor.apellidos for d in a.detalles if d.doctor]))
            
            # Calculate commission percentages
            comisiones = []
            for d in a.detalles:
                if d.porcentaje_comision > 0:
                    comisiones.append(f"{d.porcentaje_comision:.0f}%")
            
            # Calculate payment breakdown
            pagos_map = {"EF": 0, "TR": 0, "TC": 0, "AB": 0}
            for p in a.pagos:
                if p.forma_pago in pagos_map:
                    pagos_map[p.forma_pago] += p.monto

            # Calculate Global Financial Status for the Patient
            try:
                p = a.paciente
                total_consumido = sum([sum([d.total_calculado for d in at.detalles]) for at in p.atenciones])
                total_pagado_global = sum([sum([pay.monto for pay in at.pagos]) for at in p.atenciones])
                patient_balance = total_pagado_global - total_consumido
                
                patient_status = "AL_DIA"
                if patient_balance < -0.01:
                    patient_status = "DEUDA"
                elif patient_balance > 0.01:
                    patient_status = "FAVOR"
            except Exception as e_inner:
                print(f"Error calculating balance for atencion {a.id}: {e_inner}")
                patient_balance = 0
                patient_status = "ERROR"
            
            results.append({
                "id": a.id,
                "fecha": a.fecha, # datetime
                "hora": a.fecha.strftime("%H:%M"),
                "fecha_dia": a.fecha.strftime("%d de %B de %Y"), # Grouping key
                "paciente": {
                    "nombres": a.paciente.nombres,
                    "apellidos": a.paciente.apellidos,
                    "historia_clinica": a.paciente.historia_clinica
                },
                "resumen_tratamientos": ", ".join(tratamientos_names) if tratamientos_names else "-",
                "resumen_doctores": ", ".join(doctors_names) if doctors_names else "-",
                "resumen_comisiones": ", ".join(comisiones) if comisiones else "-",
                # Note: Continue mapping
                "total": a.total_atencion_valor,
                "pagos": pagos_map,
                "pagado": sum(pagos_map.values()),
                "saldo_pendiente": a.total_atencion_valor - sum(pagos_map.values()), # Re-calc locally
                "validado": a.validado,
                "estado": a.estado,
                "patient_balance": patient_balance,
                "patient_financial_status": patient_status
            })
        return results
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
@app.get("/api/atenciones/historial")
def get_historial_atenciones(
    q: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    tratamiento_id: Optional[int] = None,
    page: int = 1,
    size: int = 50,
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user)
):
    try:
        skip = (page - 1) * size
        query = (
            select(Atencion)
            .where(Atencion.sucursal_id == user.sucursal_id)
            .options(
                selectinload(Atencion.detalles).selectinload(AtencionDetalle.tratamiento),
                selectinload(Atencion.detalles).selectinload(AtencionDetalle.doctor),
                selectinload(Atencion.paciente),
                selectinload(Atencion.pagos)
            )
        )

        # 1. Filter by Doctor if applicable
        if user.role == "doctor" and user.doctor_id:
            query = query.join(Atencion.detalles, isouter=True).where(
                (Atencion.doctor_id == user.doctor_id) |
                (AtencionDetalle.doctor_id == user.doctor_id)
            ).distinct()

        # 2. Search by Patient Name
        if q:
            query = query.join(Atencion.paciente).where(
                (Paciente.nombres.contains(q)) | (Paciente.apellidos.contains(q))
            )

        # 3. Date range filter
        if start_date:
            query = query.where(Atencion.fecha >= datetime.strptime(start_date, "%Y-%m-%d"))
        if end_date:
            query = query.where(Atencion.fecha < datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1))

        # 4. Treatment filter (subquery to avoid join conflicts)
        if tratamiento_id:
            query = query.where(
                Atencion.id.in_(
                    select(AtencionDetalle.atencion_id).where(AtencionDetalle.tratamiento_id == tratamiento_id)
                )
            )
            
        # 3. Execution (With Pagination)
        atenciones = session.exec(query.order_by(Atencion.fecha.desc()).offset(skip).limit(size)).all()
        
        results = []
        for a in atenciones:
            tratamientos_names = [d.tratamiento.nombre for d in a.detalles if d.tratamiento]
            doctors_names = list(set([d.doctor.nombres + " " + d.doctor.apellidos for d in a.detalles if d.doctor]))
            
            comisiones = []
            for d in a.detalles:
                if d.porcentaje_comision > 0:
                    comisiones.append(f"{d.porcentaje_comision:.0f}%")
            
            pagos_map = {"EF": 0, "TR": 0, "TC": 0, "AB": 0}
            for p in a.pagos:
                if p.forma_pago in pagos_map:
                    pagos_map[p.forma_pago] += p.monto

            # Patient status
            patient_balance = 0.0
            patient_status = "AL_DIA"
            if a.paciente:
                # Use saldo_favor as the base for the patient's balance status
                patient_balance = float(a.paciente.saldo_favor)
                if patient_balance < -0.01:
                    patient_status = "DEUDA"
                elif patient_balance > 0.01:
                    patient_status = "FAVOR"
            
            results.append({
                "id": a.id,
                "fecha": a.fecha,
                "hora": a.fecha.strftime("%H:%M"),
                "fecha_dia": a.fecha.strftime("%d de %B de %Y"),
                "paciente": {
                    "nombres": a.paciente.nombres if a.paciente else "N/A",
                    "apellidos": a.paciente.apellidos if a.paciente else "",
                    "historia_clinica": a.paciente.historia_clinica if a.paciente else ""
                },
                "resumen_tratamientos": ", ".join(tratamientos_names) if tratamientos_names else "-",
                "resumen_doctores": ", ".join(doctors_names) if doctors_names else "-",
                "resumen_comisiones": ", ".join(comisiones) if comisiones else "-",
                "total": float(a.total_atencion_valor),
                "pagos": pagos_map,
                "pagado": sum(pagos_map.values()),
                "saldo_pendiente": float(a.total_atencion_valor - sum(pagos_map.values())),
                "validado": a.validado,
                "estado": a.estado,
                "patient_balance": patient_balance,
                "patient_financial_status": patient_status
            })
        return results
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/atenciones/{atencion_id}")
def get_atencion_detail(atencion_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    # Use eager loading to avoid lazy-load issues in serialization
    atencion = session.exec(
        select(Atencion)
        .where(Atencion.id == atencion_id)
        .options(
            selectinload(Atencion.paciente),
            selectinload(Atencion.pagos),
            selectinload(Atencion.detalles).selectinload(AtencionDetalle.tratamiento),
            selectinload(Atencion.detalles).selectinload(AtencionDetalle.doctor)
        )
    ).first()

    if not atencion:
        raise HTTPException(status_code=404, detail="Atención no encontrada")
    
    # Calculate Global Financial Status for the Patient
    p = atencion.paciente
    total_consumido_global = sum([sum([d.total_calculado for d in at.detalles]) for at in p.atenciones])
    total_pagado_global = sum([sum([pay.monto for pay in at.pagos]) for at in p.atenciones])
    patient_balance = total_pagado_global - total_consumido_global
    
    patient_status = "AL_DIA"
    if patient_balance < -0.01:
        patient_status = "DEUDA"
    elif patient_balance > 0.01:
        patient_status = "FAVOR"

    # Custom serialization ensuring clean data and unified property names
    return {
        "id": atencion.id,
        "fecha": atencion.fecha,
        "estado": atencion.estado,
        "validado": atencion.validado,
        "observaciones": atencion.observaciones,
        # Unified total names for FE consistency
        "total_atencion": atencion.total_atencion_valor,
        "total_pagado": atencion.total_pagado,
        "saldo_pendiente": atencion.saldo_pendiente,
        "paciente": {
            "id": atencion.paciente.id,
            "nombres": atencion.paciente.nombres,
            "apellidos": atencion.paciente.apellidos,
            "numero_identificacion": atencion.paciente.numero_identificacion,
            "historia_clinica": atencion.paciente.historia_clinica,
            "balance": patient_balance,
            "saldo_favor": atencion.paciente.saldo_favor,
            "financial_status": patient_status
        },
        "detalles": [
            {
                "id": d.id,
                "tratamiento_id": d.tratamiento_id,
                "doctor_id": d.doctor_id,
                "vendedor_id": d.vendedor_id,
                "tratamiento_nombre": d.tratamiento.nombre if d.tratamiento else "Desconocido",
                "get_doctor": (
                    f"Dr. {d.doctor.nombres} {d.doctor.apellidos}" if d.doctor
                    else (f"Venta: {d.vendedor.username}" if d.vendedor else "Recepción")
                ),
                "doctor_nombre": f"{d.doctor.nombres} {d.doctor.apellidos}" if d.doctor else None,
                "vendedor_nombre": d.vendedor.username if d.vendedor else None,
                "nombre_tratamiento": d.tratamiento.nombre if d.tratamiento else "Desconocido",
                "precio_unitario": d.precio_unitario,
                "cantidad": d.cantidad,
                "porcentaje_comision": d.porcentaje_comision,
                "total": d.total_calculado
            } for d in atencion.detalles
        ],
        "pagos": [
            {
                "id": p.id,
                "fecha": p.fecha,
                "forma_pago": p.forma_pago,
                "monto": p.monto
            } for p in atencion.pagos
        ],
        "totales": {
            "total_atencion": atencion.total_atencion_valor,
            "total_pagado": atencion.total_pagado,
            "saldo_pendiente": atencion.saldo_pendiente
        }
    }

@app.post("/api/atenciones/{atencion_id}/detalles")
def add_detalle(
    atencion_id: int,
    tratamiento_id: int,
    precio: float,
    cantidad: int = 1,
    doctor_id: Optional[int] = None,
    vendedor_id: Optional[int] = None,
    comision: float = 0,
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user)
):
    atencion = session.get(Atencion, atencion_id)
    if not atencion:
        raise HTTPException(status_code=404, detail="Atención no encontrada")
    if atencion.validado:
        raise HTTPException(status_code=400, detail="No se puede modificar una atención validada")
    check_recepcion_time_limit(atencion, user)

    detalle = AtencionDetalle(
        atencion_id=atencion_id,
        tratamiento_id=tratamiento_id,
        precio_unitario=precio,
        cantidad=max(1, int(cantidad)),
        porcentaje_comision=comision,
        doctor_id=doctor_id,
        vendedor_id=vendedor_id
    )
    session.add(detalle)
    
    # LOG
    t_nombre = session.get(Tratamiento, tratamiento_id).nombre
    registrar_log(atencion_id, "Tratamiento Añadido", f"Se agregó '{t_nombre}' por ${precio}", session, user)
    
    session.commit()
    return {"message": "Detalle agregado"}

@app.delete("/api/detalles/{detalle_id}")
def delete_detalle(
    detalle_id: int,
    accion: str = "dejar_abono",
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user)
):
    if user.role not in ("admin", "recepcion"):
        raise HTTPException(status_code=403, detail="No tienes permiso para eliminar tratamientos")
    detalle = session.get(AtencionDetalle, detalle_id)
    if not detalle:
        return {"message": "Detalle no encontrado"}

    atencion = session.get(Atencion, detalle.atencion_id)
    if atencion and atencion.validado:
        raise HTTPException(status_code=400, detail="No se puede modificar una atención validada")
    if atencion:
        check_recepcion_time_limit(atencion, user)

    monto_detalle = detalle.total_calculado
    nuevo_total = atencion.total_atencion_valor - monto_detalle
    exceso = atencion.total_pagado - nuevo_total

    if exceso > 0:
        if accion == "eliminar_pago":
            pagos = sorted(atencion.pagos, key=lambda p: p.fecha, reverse=True)
            por_reducir = Decimal(str(exceso))
            for pago in pagos:
                if por_reducir <= 0:
                    break
                if pago.monto <= por_reducir:
                    por_reducir -= pago.monto
                    session.delete(pago)
                else:
                    pago.monto -= por_reducir
                    por_reducir = Decimal("0")
                    session.add(pago)
        else:  # dejar_abono
            paciente = session.get(Paciente, atencion.paciente_id)
            if paciente:
                paciente.saldo_favor += Decimal(str(exceso))
                session.add(paciente)
                abono = HistorialAbono(
                    paciente_id=paciente.id,
                    usuario_id=user.id,
                    monto=Decimal(str(exceso)),
                    metodo_pago="ABONO"
                )
                session.add(abono)

    registrar_log(detalle.atencion_id, "Tratamiento Eliminado", f"Se eliminó '{detalle.tratamiento.nombre}'", session)
    session.delete(detalle)
    session.commit()
    return {"message": "Detalle eliminado"}

from pydantic import BaseModel

class UpdateDetail(BaseModel):
    tratamiento_id: int
    doctor_id: Optional[int] = None
    vendedor_id: Optional[int] = None
    cantidad: int = 1
    precio: float
    comision: float

@app.put("/api/detalles/{detalle_id}")
def update_detalle(detalle_id: int, data: UpdateDetail, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    detalle = session.get(AtencionDetalle, detalle_id)
    if not detalle:
        raise HTTPException(status_code=404, detail="Detalle no encontrado")

    atencion = session.get(Atencion, detalle.atencion_id)
    if atencion and atencion.validado:
        raise HTTPException(status_code=400, detail="No se puede modificar una atención validada")
    if atencion:
        check_recepcion_time_limit(atencion, user)

    detalle.tratamiento_id = data.tratamiento_id
    detalle.cantidad = max(1, int(data.cantidad))
    detalle.doctor_id = data.doctor_id
    
    # If the user sets a doctor, wipe the vendedor_id.
    # Otherwise, preserve the existing vendedor_id or assign it to the updater if it's a new 'reception sale' assignment.
    if data.doctor_id:
         detalle.vendedor_id = None
    elif data.vendedor_id is not None:
         detalle.vendedor_id = data.vendedor_id
    elif detalle.vendedor_id is None and not data.doctor_id:
         # If they explicitly chose 'recepcion' (doctor_id is None) and it had NO vendedor before, give it to them
         user = session.exec(select(User).where(User.username == "admin")).first() # Can't get current user easily here without dependency, but let's just leave it as is if they just updated price.
         
    detalle.precio_unitario = data.precio
    detalle.porcentaje_comision = data.comision

    
    session.add(detalle)
    session.commit()
    return {"message": "Detalle actualizado"}

class UpdateAtencion(BaseModel):
    observaciones: str | None = None

@app.put("/api/atenciones/{atencion_id}/fecha")
def update_atencion_fecha(atencion_id: int, fecha: str, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Solo el administrador puede cambiar la fecha de una atención")
    atencion = session.get(Atencion, atencion_id)
    if not atencion:
        raise HTTPException(status_code=404, detail="Atención no encontrada")
    try:
        atencion.fecha = datetime.fromisoformat(fecha)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    registrar_log(atencion_id, "Fecha Modificada", f"Fecha cambiada a {fecha}", session, user)
    session.add(atencion)
    session.commit()
    return {"message": "Fecha actualizada"}

@app.put("/api/atenciones/{atencion_id}")
def update_atencion(atencion_id: int, data: UpdateAtencion, session: Session = Depends(get_session)):
    atencion = session.get(Atencion, atencion_id)
    if not atencion:
        raise HTTPException(status_code=404, detail="Atención no encontrada")
        
    if atencion.validado:
        raise HTTPException(status_code=400, detail="No se puede modificar una atención validada")

    if data.observaciones is not None:
        atencion.observaciones = data.observaciones
        registrar_log(atencion_id, "Notas Actualizadas", "Se actualizaron las notas clínicas/evolución", session)
    
    session.add(atencion)
    session.commit()
    return {"message": "Atención actualizada"}

from pydantic import BaseModel
class PaymentSync(BaseModel):
    efectivo: float
    transferencia: float
    tarjeta: float
    abono: float
    metodo_excedente: Optional[str] = None
    confirmar_abono: bool = False

@app.post("/api/atenciones/{atencion_id}/pagos/sync")
def sync_pagos(atencion_id: int, data: PaymentSync, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    atencion = session.get(Atencion, atencion_id)
    if not atencion:
        raise HTTPException(status_code=404, detail="Atención no encontrada")
    if atencion.validado:
        raise HTTPException(status_code=400, detail="No se puede modificar una atención validada")

    pago_directo = Decimal(str(data.efectivo)) + Decimal(str(data.transferencia)) + Decimal(str(data.tarjeta))
    if pago_directo > 0 and len(atencion.detalles) == 0 and not data.confirmar_abono:
        raise HTTPException(
            status_code=409,
            detail={"code": "SIN_TRATAMIENTOS", "monto": float(pago_directo)}
        )

    # --- WALLET ADJUSTMENT (Inmediata) ---
    # Calculamos la diferencia entre el abono anterior y el nuevo
    old_abono = Decimal(sum([p.monto for p in atencion.pagos if p.forma_pago == "AB"]))
    # Usar str() para evitar problemas de precisión de float a Decimal
    new_abono = Decimal(str(data.abono)) 
    diff = new_abono - old_abono

    if diff != 0:
        if diff > 0 and atencion.paciente.saldo_favor < diff:
             raise HTTPException(status_code=400, detail=f"Saldo insuficiente en billetera. Disponible: ${atencion.paciente.saldo_favor}, Requerido adicional: ${diff}")
        
        atencion.paciente.saldo_favor -= diff
        session.add(atencion.paciente)
        registrar_log(atencion_id, "Billetera Ajustada", f"Uso de billetera cambiado de ${old_abono} a ${new_abono}", session, user)

    # --- OVERPAYMENT LOGIC (Automatic Recharge) ---
    new_efectivo = Decimal(str(data.efectivo))
    new_transferencia = Decimal(str(data.transferencia))
    new_tarjeta = Decimal(str(data.tarjeta))

    total_received = new_efectivo + new_transferencia + new_tarjeta + new_abono
    total_due = atencion.total_atencion_valor

    # Revert any previous auto-surplus for this atencion before recalculating
    prev_historial = session.exec(
        select(HistorialAbono)
        .where(HistorialAbono.atencion_id == atencion_id)
    ).all()
    for ph in prev_historial:
        atencion.paciente.saldo_favor -= ph.monto
        session.delete(ph)
    session.add(atencion.paciente)

    if total_received > total_due:
        surplus = total_received - total_due

        # 1. Credit surplus to Wallet
        atencion.paciente.saldo_favor += surplus
        session.add(atencion.paciente)

        metodo_final = data.metodo_excedente if data.metodo_excedente else "Desconocido"
        historial = HistorialAbono(
            paciente_id=atencion.paciente.id,
            usuario_id=user.id if user else None,
            atencion_id=atencion_id,
            monto=surplus,
            metodo_pago=metodo_final,
            fecha=atencion.fecha,
        )
        session.add(historial)

        registrar_log(atencion_id, "Recarga Automática", f"Sobrepago de ${surplus} acreditado a Billetera", session, user)

        # 2. Adjust payments to not exceed total_due for accounting purposes
        remaining_reduction = surplus
        
        # Helper to safely reduce a money metric and the remaining reduction
        def reduce_amount(current_amount, target_reduction):
             deduct = min(current_amount, target_reduction)
             return current_amount - deduct, target_reduction - deduct

        # Prioritize deducting from the method the user explicitly declared as the surplus source
        if metodo_final == "Efectivo" and remaining_reduction > 0 and new_efectivo > 0:
             new_efectivo, remaining_reduction = reduce_amount(new_efectivo, remaining_reduction)
             
        if metodo_final == "Transferencia" and remaining_reduction > 0 and new_transferencia > 0:
             new_transferencia, remaining_reduction = reduce_amount(new_transferencia, remaining_reduction)
             
        if metodo_final == "Tarjeta" and remaining_reduction > 0 and new_tarjeta > 0:
             new_tarjeta, remaining_reduction = reduce_amount(new_tarjeta, remaining_reduction)
             
        # If there is STILL remaining reduction (e.g. they typed TR: 10, EF: 120 and chose TR as source),
        # fallback to waterfall reduction to guarantee exact total_due matches.
        if remaining_reduction > 0 and new_efectivo > 0:
             new_efectivo, remaining_reduction = reduce_amount(new_efectivo, remaining_reduction)
             
        if remaining_reduction > 0 and new_transferencia > 0:
             new_transferencia, remaining_reduction = reduce_amount(new_transferencia, remaining_reduction)
             
        if remaining_reduction > 0 and new_tarjeta > 0:
             new_tarjeta, remaining_reduction = reduce_amount(new_tarjeta, remaining_reduction)

    # Informative log about final payments (adjusted)
    registrar_log(atencion_id, "Pagos Sincronizados", f"Finales - Ef: ${new_efectivo}, Tr: ${new_transferencia}, Tc: ${new_tarjeta}", session, user)

    # Wipe existing payments
    for p in atencion.pagos:
        session.delete(p)
    
    # Add new payments (using adjusted values)
    pago_fecha = atencion.fecha
    if new_efectivo > 0:
        session.add(Pago(atencion_id=atencion_id, forma_pago="EF", monto=new_efectivo, fecha=pago_fecha))
    if new_transferencia > 0:
        session.add(Pago(atencion_id=atencion_id, forma_pago="TR", monto=new_transferencia, fecha=pago_fecha))
    if new_tarjeta > 0:
        session.add(Pago(atencion_id=atencion_id, forma_pago="TC", monto=new_tarjeta, fecha=pago_fecha))
    if new_abono > 0:
        session.add(Pago(atencion_id=atencion_id, forma_pago="AB", monto=new_abono, fecha=pago_fecha))

    session.commit()
    return {"message": "Pagos actualizados"}

@app.put("/api/atenciones/{atencion_id}/pagos")
def update_atencion_pagos(atencion_id: int, data: dict, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    """
    Specifically for Secretary to adjust payments after attention is finished.
    Expects data: {"efectivo": X, "transferencia": Y, "tarjeta": Z, "abono": W}
    """
    atencion = session.get(Atencion, atencion_id)
    if not atencion:
        raise HTTPException(status_code=404, detail="Atención no encontrada")
    
    # Validation: Only secretary/admin can use this specific post-closure edit
    if user.role not in ["admin", "recepcion"]:
        raise HTTPException(status_code=403, detail="No tiene permisos para editar pagos")

    # --- WALLET ADJUSTMENT (Inmediata) ---
    old_abono = sum([p.monto for p in atencion.pagos if p.forma_pago == "AB"])
    new_abono = Decimal(data.get("abono", 0))
    diff = new_abono - Decimal(old_abono)

    if diff != 0:
        if diff > 0 and atencion.paciente.saldo_favor < diff:
             raise HTTPException(status_code=400, detail=f"Saldo insuficiente en billetera.")
        
        atencion.paciente.saldo_favor -= diff
        session.add(atencion.paciente)
        registrar_log(atencion_id, "Billetera Ajustada (Manual)", f"Abono cambiado de ${old_abono} a ${new_abono}", session, user)

    # --- OVERPAYMENT LOGIC (Automatic Recharge) ---
    new_efectivo = Decimal(data.get("efectivo", 0))
    new_transferencia = Decimal(data.get("transferencia", 0))
    new_tarjeta = Decimal(data.get("tarjeta", 0))
    
    total_received = new_efectivo + new_transferencia + new_tarjeta + new_abono
    total_due = atencion.total_atencion_valor

    if total_received > total_due:
        surplus = total_received - total_due
        
        # 1. Credit surplus to Wallet
        atencion.paciente.saldo_favor += surplus
        session.add(atencion.paciente)
        registrar_log(atencion_id, "Recarga Automática", f"Sobrepago de ${surplus} acreditado a Billetera", session, user)

        # 2. Adjust payments to not exceed total_due for accounting purposes
        # Strategy: Reduce from Cash first, then Transfer, then Card
        remaining_reduction = surplus
        
        if remaining_reduction > 0 and new_efectivo > 0:
            deduct = min(new_efectivo, remaining_reduction)
            new_efectivo -= deduct
            remaining_reduction -= deduct
            
        if remaining_reduction > 0 and new_transferencia > 0:
            deduct = min(new_transferencia, remaining_reduction)
            new_transferencia -= deduct
            remaining_reduction -= deduct
            
        if remaining_reduction > 0 and new_tarjeta > 0:
            deduct = min(new_tarjeta, remaining_reduction)
            new_tarjeta -= deduct
            remaining_reduction -= deduct

    # Informative log
    registrar_log(atencion_id, "Pagos Editados (Manual)", f"Nuevos montos - Ef: {new_efectivo}, Tr: {new_transferencia}, Tc: {new_tarjeta}", session, user)

    # Wipe old payments for this attention
    for p in atencion.pagos:
        session.delete(p)
    session.flush() # Ensure deletions are processed before adding new ones
    
    # Add new payments (use adjusted values)
    pago_fecha = atencion.fecha
    if new_efectivo > 0:
        session.add(Pago(atencion_id=atencion_id, forma_pago="EF", monto=new_efectivo, fecha=pago_fecha))
    if new_transferencia > 0:
        session.add(Pago(atencion_id=atencion_id, forma_pago="TR", monto=new_transferencia, fecha=pago_fecha))
    if new_tarjeta > 0:
        session.add(Pago(atencion_id=atencion_id, forma_pago="TC", monto=new_tarjeta, fecha=pago_fecha))
    if new_abono > 0:
        session.add(Pago(atencion_id=atencion_id, forma_pago="AB", monto=new_abono, fecha=pago_fecha))

    session.commit()
    return {"message": "Pagos actualizados correctamente"}


@app.post("/api/atenciones/{atencion_id}/pagos/add")
def add_atencion_pago(atencion_id: int, data: dict, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    """
    Transactional Add Payment.
    Expects data: {"forma_pago": "EF"|"TR"|"TC"|"AB", "monto": 10.00}
    """
    atencion = session.get(Atencion, atencion_id)
    if not atencion:
        raise HTTPException(status_code=404, detail="Atención no encontrada")
    
    if user.role not in ["admin", "recepcion"]:
        raise HTTPException(status_code=403, detail="No tiene permisos para agregar pagos")

    forma_pago = data.get("forma_pago")
    if forma_pago not in ["EF", "TR", "TC", "AB"]:
        raise HTTPException(status_code=400, detail="Forma de pago inválida")

    try:
        monto = Decimal(data.get("monto", 0))
    except:
         raise HTTPException(status_code=400, detail="Monto inválido")
         
    if monto <= 0:
        raise HTTPException(status_code=400, detail="El monto debe ser positivo")

    # --- ABONO LOGIC (Wallet Deduction) ---
    if forma_pago == "AB":
        if atencion.paciente.saldo_favor < monto:
             raise HTTPException(status_code=400, detail=f"Saldo insuficiente en billetera (${atencion.paciente.saldo_favor})")
        
        # Deduct from wallet immediately
        atencion.paciente.saldo_favor -= monto
        session.add(atencion.paciente)
        registrar_log(atencion_id, "Pago con Billetera", f"Se descontaron ${monto} de la billetera", session, user)

    # --- REGULAR PAYMENTS & SURPLUS LOGIC ---
    else:
        # Check for overpayment against the TOTAL DUE vs (TOTAL PAID + CURRENT AMT)
        total_due = atencion.total_atencion_valor
        total_paid = atencion.total_pagado # Logic in model sums existing payments
        
        if total_paid + monto > total_due:
            surplus = (total_paid + monto) - total_due
            
            # If the surplus is the ENTIRE amount (already paid in full), credit all to wallet
            # If partial, credit only the excess
            
            credit_to_wallet = surplus
            payment_record = monto - surplus # This ensures the attention balances exactly to 0 pending
            
            if credit_to_wallet > 0:
                atencion.paciente.saldo_favor += credit_to_wallet
                session.add(atencion.paciente)

                # NEW: Record the overpayment as a proper Wallet Recharge in HistorialAbono
                metodo_excedente = data.get("metodo_excedente", forma_pago)
                
                historial = HistorialAbono(
                    paciente_id=atencion.paciente.id,
                    usuario_id=user.id if user else None,
                    monto=Decimal(credit_to_wallet),
                    metodo_pago=metodo_excedente,
                    fecha=atencion.fecha,
                )
                session.add(historial)

                registrar_log(atencion_id, "Sobrepago a Billetera", f"Excedente de ${credit_to_wallet} acreditado a billetera", session, user)
            
            # If there is a remaining valid payment part, update 'monto' to that
            # BUT: Transactions normally record the actual event. 
            # If I hand you $100 bill for a $50 debt. You record "Paid $50, changed $50".
            # The system wants to record payments attached to the attention. 
            # If we record $100 payment, the balance becomes -$50.
            # To strictly balance the attention, we record $50 payment. 
            
            monto = payment_record 
            
            if monto <= 0 and credit_to_wallet > 0:
                 # Fully covered by overpayment logic (e.g. debt was 0)
                 # We just return success as we moved money to wallet
                 session.commit()
                 return {"message": "Monto acreditado totalmente a billetera"}

    # Record the payment
    if monto > 0:
        pago = Pago(atencion_id=atencion_id, forma_pago=forma_pago, monto=monto)
        session.add(pago)
        registrar_log(atencion_id, "Pago Agregado", f"Pago de ${monto} vía {forma_pago}", session, user)
    
    session.commit()
    return {"message": "Pago registrado correctamente", "pagos": atencion.pagos}

@app.delete("/api/atenciones/{atencion_id}")
def delete_atencion(atencion_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    if user.role not in ("admin", "recepcion"):
        raise HTTPException(status_code=403, detail="Solo admin o recepción pueden eliminar atenciones")
    atencion = session.get(Atencion, atencion_id)
    if not atencion:
        raise HTTPException(status_code=404, detail="Atención no encontrada")
    if atencion.validado:
        raise HTTPException(status_code=400, detail="No se puede eliminar una atención validada")
    check_recepcion_time_limit(atencion, user)

    paciente = session.exec(
        select(Paciente)
        .where(Paciente.id == atencion.paciente_id)
    ).first()

    if paciente:
        # 1. Revertir sobrepagos que se acreditaron a la billetera desde esta atencion
        historial_atencion = session.exec(
            select(HistorialAbono).where(HistorialAbono.atencion_id == atencion_id)
        ).all()
        for h in historial_atencion:
            paciente.saldo_favor -= h.monto
            session.delete(h)

        # 2. Devolver montos pagados con billetera (AB) al saldo del paciente
        total_abono = sum([p.monto for p in atencion.pagos if p.forma_pago == "AB"])
        if total_abono > 0:
            paciente.saldo_favor += total_abono

        session.add(paciente)

    session.delete(atencion)
    session.commit()
    return {"message": "Atención eliminada"}

@app.get("/api/pagos/bonos")
def get_bonos_report(
    start_date: str = None, 
    end_date: str = None, 
    search: str = None, 
    page: int = 1,
    size: int = 50,
    session: Session = Depends(get_session), 
    user: User = Depends(get_current_user)
):
    # Basic security check
    if user.role not in ["admin", "recepcion"]:
        raise HTTPException(status_code=403, detail="No tiene permisos para ver reportes de pagos")

    skip = (page - 1) * size
    query = (
        select(HistorialAbono)
        .join(Paciente) # Ensure connection to Paciente
    )

    if search:
        search_val = f"%{search}%"
        query = query.where(
            (Paciente.nombres.ilike(search_val)) | 
            (Paciente.apellidos.ilike(search_val)) | 
            (Paciente.numero_identificacion.ilike(search_val))
        )

    query = query.order_by(HistorialAbono.fecha.desc()).offset(skip).limit(size)

    # Date filters (optional)
    if start_date:
        query = query.where(HistorialAbono.fecha >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        # End date inclusive
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        query = query.where(HistorialAbono.fecha < end_dt)

    bonos = session.exec(query).all()
    
    results = []
    for b in bonos:
        results.append({
            "id": b.id,
            "fecha": b.fecha,
            "paciente": f"{b.paciente.nombres} {b.paciente.apellidos}",
            "monto": b.monto,
            "metodo_pago": b.metodo_pago,
            "usuario": b.usuario.username if b.usuario else "Desconocido"
        })
        
    return results


@app.get("/api/users/me")
def read_users_me(current_user: User = Depends(get_current_user)):
    return current_user

@app.post("/api/atenciones/{atencion_id}/validar")
def validar_atencion(atencion_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    # Eager load relationships for inventory logic
    atencion = session.exec(
        select(Atencion)
        .where(Atencion.id == atencion_id)
        .options(
            selectinload(Atencion.detalles).selectinload(AtencionDetalle.tratamiento).selectinload(Tratamiento.insumos_requeridos).selectinload(Receta.insumo)
        )
    ).first()
    
    if atencion:
        if atencion.validado:
             return {"message": "Ya estaba validado"}
             
        atencion.validado = True
        atencion.estado = "FINALIZADO"
        
        # --- INVENTORY DEDUCTION (Manual) ---
        process_stock_deduction(atencion, session)

        # --- WALLET DEDUCTION ---
        # Removido: Ahora el descuento ocurre en tiempo real durante la edición (sync_pagos)

        session.add(atencion)
        session.commit()
    return {"message": "Validado con éxito"}

@app.get("/api/reportes/reporte-doctores")
def reporte_doctores(
    start_date: str = None,
    end_date: str = None,
    doctor_id: Optional[int] = None,
    tratamiento_id: Optional[int] = None,
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user)
):
    """
    Detailed doctor commission report.
    Returns one row per AtencionDetalle with:
    fecha, doctor, tratamiento, valor_cancelado (treatment value),
    pct_comision, valor_comision (theoretical), valor_comision_pagada, comision_pendiente.
    """
    if user.role not in ["admin", "recepcion"]:
        raise HTTPException(status_code=403, detail="No tiene permisos")

    q = (
        select(AtencionDetalle)
        .join(Atencion)
        .where(Atencion.sucursal_id == user.sucursal_id)
        .options(
            selectinload(AtencionDetalle.doctor),
            selectinload(AtencionDetalle.tratamiento),
            selectinload(AtencionDetalle.atencion)
        )
        .order_by(Atencion.fecha.asc())
    )
    if start_date:
        q = q.where(Atencion.fecha >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        q = q.where(Atencion.fecha < end_dt)
    if doctor_id:
        q = q.where(AtencionDetalle.doctor_id == doctor_id)
    if tratamiento_id:
        q = q.where(AtencionDetalle.tratamiento_id == tratamiento_id)

    detalles = session.exec(q).all()

    filas = []
    for d in detalles:
        valor_trat   = float(d.total_calculado)
        pct          = float(d.porcentaje_comision)
        val_comision = round(valor_trat * pct / 100, 2)
        val_pagado   = round(float(d.comision_pagada_monto or 0), 2)
        pendiente    = round(val_comision - val_pagado, 2)
        filas.append({
            "fecha":               d.atencion.fecha.strftime("%Y-%m-%d") if d.atencion else "",
            "doctor":              f"{d.doctor.nombres} {d.doctor.apellidos}" if d.doctor else "—",
            "tratamiento":         d.tratamiento.nombre if d.tratamiento else "—",
            "valor_cancelado":     valor_trat,
            "pct_comision":        pct,
            "valor_comision":      val_comision,
            "valor_comision_pagada": val_pagado,
            "comision_pendiente":  max(pendiente, 0)
        })

    totales = {
        "valor_cancelado":       round(sum(f["valor_cancelado"]       for f in filas), 2),
        "valor_comision":        round(sum(f["valor_comision"]        for f in filas), 2),
        "valor_comision_pagada": round(sum(f["valor_comision_pagada"] for f in filas), 2),
        "comision_pendiente":    round(sum(f["comision_pendiente"]    for f in filas), 2),
    }

    return {"filas": filas, "totales": totales}

@app.get("/api/reportes/tratamientos-en-curso")
def reporte_tratamientos_en_curso(
    start_date: str = None,
    end_date: str = None,
    tratamiento_id: Optional[int] = None,
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user)
):
    """
    Two blocks:
    1. Patients who STARTED a treatment within the date range (new starters).
    2. Patients currently ACTIVE in a treatment (regardless of date).
    Both filterable by tratamiento.
    """
    if user.role not in ["admin", "recepcion"]:
        raise HTTPException(status_code=403, detail="No tiene permisos")

    # --- Block 1: new starters in period ---
    q_new = (
        select(TratamientoEnCurso)
        .options(
            selectinload(TratamientoEnCurso.paciente),
            selectinload(TratamientoEnCurso.tratamiento)
        )
        .order_by(TratamientoEnCurso.fecha_inicio.asc())
    )
    if start_date:
        q_new = q_new.where(TratamientoEnCurso.fecha_inicio >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        q_new = q_new.where(TratamientoEnCurso.fecha_inicio < end_dt)
    if tratamiento_id:
        q_new = q_new.where(TratamientoEnCurso.tratamiento_id == tratamiento_id)

    nuevos = session.exec(q_new).all()
    nuevos_list = [
        {
            "id": r.id,
            "paciente": r.paciente.nombre_mostrar if r.paciente else "—",
            "historia_clinica": r.paciente.historia_clinica if r.paciente else "—",
            "tratamiento": r.tratamiento.nombre if r.tratamiento else "—",
            "fecha_inicio": r.fecha_inicio.strftime("%Y-%m-%d"),
            "activo": r.activo,
            "fecha_fin": r.fecha_fin.strftime("%Y-%m-%d") if r.fecha_fin else None,
        }
        for r in nuevos
    ]

    # --- Block 2: currently active (no date filter, only tratamiento filter) ---
    q_activos = (
        select(TratamientoEnCurso)
        .where(TratamientoEnCurso.activo == True)
        .options(
            selectinload(TratamientoEnCurso.paciente),
            selectinload(TratamientoEnCurso.tratamiento)
        )
        .order_by(TratamientoEnCurso.tratamiento_id, TratamientoEnCurso.fecha_inicio.asc())
    )
    if tratamiento_id:
        q_activos = q_activos.where(TratamientoEnCurso.tratamiento_id == tratamiento_id)

    activos = session.exec(q_activos).all()

    # Group by tratamiento for summary
    resumen_activos: dict = {}
    activos_list = []
    for r in activos:
        t_nombre = r.tratamiento.nombre if r.tratamiento else "—"
        resumen_activos[t_nombre] = resumen_activos.get(t_nombre, 0) + 1
        activos_list.append({
            "id": r.id,
            "paciente": r.paciente.nombre_mostrar if r.paciente else "—",
            "historia_clinica": r.paciente.historia_clinica if r.paciente else "—",
            "tratamiento": t_nombre,
            "fecha_inicio": r.fecha_inicio.strftime("%Y-%m-%d"),
        })

    resumen_list = [{"tratamiento": k, "total": v} for k, v in sorted(resumen_activos.items())]

    return {
        "nuevos": nuevos_list,
        "activos": activos_list,
        "resumen_activos": resumen_list,
        "total_nuevos": len(nuevos_list),
        "total_activos": len(activos_list),
    }

@app.get("/api/reportes/pacientes-por-tratamiento")
def reporte_pacientes_tratamiento(session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    # Reporte Real: Pacientes Activos en Tratamiento
    results = session.exec(
        select(Tratamiento.nombre, func.count(TratamientoEnCurso.id))
        .join(TratamientoEnCurso, Tratamiento.id == TratamientoEnCurso.tratamiento_id)
        .where(TratamientoEnCurso.activo == True)
        .group_by(Tratamiento.id)
    ).all()
    
    data = [{"nombre": r[0], "total_pacientes": r[1]} for r in results]
    return data

@app.get("/api/reportes/ingresos-mensuales")
def reporte_ingresos_mensuales(start_date: str = None, end_date: str = None, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    if user.role not in ["admin", "recepcion"]:
        raise HTTPException(status_code=403, detail="No tiene permisos")
        
    query_pagos = select(Pago)
    if user.sucursal_id:
        query_pagos = query_pagos.join(Atencion).where(Atencion.sucursal_id == user.sucursal_id)
        
    if start_date:
        query_pagos = query_pagos.where(Pago.fecha >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        query_pagos = query_pagos.where(Pago.fecha < end_dt)
        
    pagos = session.exec(query_pagos).all()

    query_auditoria = select(AuditoriaAtencion).where(AuditoriaAtencion.accion == "RECARGA_BILLETERA")
    if start_date:
        query_auditoria = query_auditoria.where(AuditoriaAtencion.fecha >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        query_auditoria = query_auditoria.where(AuditoriaAtencion.fecha < end_dt)
        
    recargas_logs = session.exec(query_auditoria).all()

    ingresos_por_mes = {}
    
    for p in pagos:
        mes_key = p.fecha.strftime("%Y-%m")
        if mes_key not in ingresos_por_mes:
            ingresos_por_mes[mes_key] = 0.0
        ingresos_por_mes[mes_key] += float(p.monto)
        
    import re
    for log in recargas_logs:
        mes_key = log.fecha.strftime("%Y-%m")
        match = re.search(r"\$(\d+(\.\d+)?)", log.descripcion)
        if match:
            if mes_key not in ingresos_por_mes:
                ingresos_por_mes[mes_key] = 0.0
            ingresos_por_mes[mes_key] += float(match.group(1))

    ingresos_list = [{"mes": k, "total": round(v, 2)} for k, v in sorted(ingresos_por_mes.items())]

    # Part 2: Produccion por doctor
    query_detalles = select(AtencionDetalle).join(Atencion).where(Atencion.validado == True).where(AtencionDetalle.doctor_id != None)
    if user.sucursal_id:
        query_detalles = query_detalles.where(Atencion.sucursal_id == user.sucursal_id)
        
    if start_date:
        query_detalles = query_detalles.where(Atencion.fecha >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        query_detalles = query_detalles.where(Atencion.fecha < end_dt)
        
    query_detalles = query_detalles.options(selectinload(AtencionDetalle.doctor), selectinload(AtencionDetalle.tratamiento))
    detalles = session.exec(query_detalles).all()

    doctores_prod = {}
    for d in detalles:
        doc_id = d.doctor_id
        if doc_id not in doctores_prod:
            doctores_prod[doc_id] = {
                "doctor_id": doc_id,
                "nombre": f"{d.doctor.nombres} {d.doctor.apellidos}" if d.doctor else "Desconocido",
                "total_producido": 0.0,
                "tratamientos": {}
            }
        
        val = float(d.total_calculado)
        doctores_prod[doc_id]["total_producido"] += val
        
        t_nombre = d.tratamiento.nombre if d.tratamiento else "Otro"
        if t_nombre not in doctores_prod[doc_id]["tratamientos"]:
             doctores_prod[doc_id]["tratamientos"][t_nombre] = 0.0
        doctores_prod[doc_id]["tratamientos"][t_nombre] += val

    for doc_id, prod in doctores_prod.items():
        trats = []
        for t_k, t_v in prod["tratamientos"].items():
            trats.append({"nombre": t_k, "monto": round(t_v, 2)})
        trats.sort(key=lambda x: x["monto"], reverse=True)
        prod["tratamientos"] = trats
        prod["total_producido"] = round(prod["total_producido"], 2)

    doctores_list = list(doctores_prod.values())
    doctores_list.sort(key=lambda x: x["total_producido"], reverse=True)

    return {
        "ingresos_mensuales": ingresos_list,
        "produccion_doctores": doctores_list
    }

@app.get("/api/reportes/flujo-caja-global")
def reporte_flujo_caja_global(start_date: str = None, end_date: str = None, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    """
    Returns total income (excluding 'AB' payments) and total expenses.
    """
    if user.role == "doctor":
        raise HTTPException(status_code=403, detail="No autorizado")

    # 1. Ingresos: Pagos (forma_pago != 'AB')
    query_pagos = select(Pago).join(Atencion).where(Atencion.sucursal_id == user.sucursal_id).where(Pago.forma_pago != 'AB')
    if start_date:
        query_pagos = query_pagos.where(Pago.fecha >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        query_pagos = query_pagos.where(Pago.fecha < end_dt)
        
    pagos = session.exec(query_pagos).all()
    total_ingresos = sum(float(p.monto) for p in pagos)
    
    # 2. Egresos: Gastos
    query_gastos = select(Gasto).where(Gasto.sucursal_id == user.sucursal_id)
    if start_date:
        query_gastos = query_gastos.where(Gasto.fecha >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        query_gastos = query_gastos.where(Gasto.fecha < end_dt)
        
    gastos = session.exec(query_gastos).all()
    total_egresos = sum(float(g.monto) for g in gastos)
    
    saldo_neto = total_ingresos - total_egresos
    
    return {
        "ingresos": total_ingresos,
        "egresos": total_egresos,
        "saldo_neto": saldo_neto
    }

@app.get("/api/reportes/financiero-completo")
def reporte_financiero_completo(session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    """
    Returns a comprehensive JSON with all attentions, details, and payments
    for Excel export.
    """
    if user.role not in ["admin", "recepcion"]:
        raise HTTPException(status_code=403, detail="No tiene permisos")

    # Load everything needed
    atenciones = session.exec(
        select(Atencion)
        .options(
            selectinload(Atencion.paciente),
            selectinload(Atencion.doctor),
            selectinload(Atencion.sucursal),
            selectinload(Atencion.detalles).selectinload(AtencionDetalle.tratamiento),
            selectinload(Atencion.detalles).selectinload(AtencionDetalle.doctor),
            selectinload(Atencion.pagos)
        )
        .order_by(Atencion.fecha.desc())
    ).all()

    report_results = []
    for at in atenciones:
        report_results.append({
            "id": at.id,
            "fecha": at.fecha.isoformat(),
            "estado": at.estado,
            "paciente": f"{at.paciente.nombres} {at.paciente.apellidos}" if at.paciente else "N/A",
            "paciente_id": at.paciente.numero_identificacion if at.paciente else "N/A",
            "doctor_principal": f"{at.doctor.nombres} {at.doctor.apellidos}" if at.doctor else "N/A",
            "sucursal": at.sucursal.nombre if at.sucursal else "N/A",
            "observaciones": at.observaciones,
            "detalles": [
                {
                    "tratamiento": d.tratamiento.nombre if d.tratamiento else "Desconocido",
                    "doctor_procedimiento": f"{d.doctor.nombres} {d.doctor.apellidos}" if d.doctor else "N/A",
                    "cantidad": d.cantidad,
                    "precio_unitario": float(d.precio_unitario),
                    "total": float(d.total_calculado)
                } for d in at.detalles
            ],
            "pagos": [
                {
                    "fecha": p.fecha.isoformat(),
                    "forma_pago": p.forma_pago,
                    "monto": float(p.monto),
                    "referencia": p.referencia
                } for p in at.pagos
            ]
        })

    return report_results

@app.get("/api/reportes/resumen-financiero")
def reporte_resumen_financiero(
    start_date: str = None,
    end_date: str = None,
    doctor_id: Optional[int] = None,
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user)
):
    """
    Comprehensive financial report:
    - Income breakdown by treatment type (quantity + amount)
    - Income breakdown by doctor (with treatment detail)
    - Expense breakdown by category (NÓMINA, INSUMOS, SERVICIOS BÁSICOS, etc.)
    - Balance general
    """
    if user.role not in ["admin", "recepcion"]:
        raise HTTPException(status_code=403, detail="No tiene permisos")

    start_dt = datetime.strptime(start_date, "%Y-%m-%d") if start_date else None
    end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) if end_date else None

    # 1. Load AtencionDetalles filtered by date, sucursal, optionally doctor
    q_detalles = (
        select(AtencionDetalle)
        .join(Atencion)
        .where(Atencion.sucursal_id == user.sucursal_id)
        .options(
            selectinload(AtencionDetalle.tratamiento),
            selectinload(AtencionDetalle.doctor)
        )
    )
    if start_dt:
        q_detalles = q_detalles.where(Atencion.fecha >= start_dt)
    if end_dt:
        q_detalles = q_detalles.where(Atencion.fecha < end_dt)
    if doctor_id:
        q_detalles = q_detalles.where(AtencionDetalle.doctor_id == doctor_id)

    detalles = session.exec(q_detalles).all()

    # Aggregate ingresos by treatment
    trat_map = {}
    for d in detalles:
        t_nombre = d.tratamiento.nombre if d.tratamiento else "Sin tratamiento"
        if t_nombre not in trat_map:
            trat_map[t_nombre] = {"nombre": t_nombre, "cantidad": 0, "monto_total": 0.0}
        trat_map[t_nombre]["cantidad"] += d.cantidad
        trat_map[t_nombre]["monto_total"] += float(d.total_calculado)

    ingresos_por_tratamiento = sorted(trat_map.values(), key=lambda x: x["monto_total"], reverse=True)
    for i in ingresos_por_tratamiento:
        i["monto_total"] = round(i["monto_total"], 2)
    total_produccion = round(sum(i["monto_total"] for i in ingresos_por_tratamiento), 2)

    # Aggregate ingresos by doctor
    doc_map = {}
    for d in detalles:
        if not d.doctor_id:
            continue
        doc_id = d.doctor_id
        if doc_id not in doc_map:
            doc_nombre = f"{d.doctor.nombres} {d.doctor.apellidos}" if d.doctor else "Desconocido"
            doc_map[doc_id] = {"doctor_id": doc_id, "nombre": doc_nombre, "total": 0.0, "tratamientos": {}}
        t_nombre = d.tratamiento.nombre if d.tratamiento else "Sin tratamiento"
        monto = float(d.total_calculado)
        doc_map[doc_id]["total"] += monto
        if t_nombre not in doc_map[doc_id]["tratamientos"]:
            doc_map[doc_id]["tratamientos"][t_nombre] = {"nombre": t_nombre, "cantidad": 0, "monto": 0.0}
        doc_map[doc_id]["tratamientos"][t_nombre]["cantidad"] += d.cantidad
        doc_map[doc_id]["tratamientos"][t_nombre]["monto"] += monto

    ingresos_por_doctor = []
    for doc_id, data in doc_map.items():
        trats = sorted(data["tratamientos"].values(), key=lambda x: x["monto"], reverse=True)
        for t in trats:
            t["monto"] = round(t["monto"], 2)
        ingresos_por_doctor.append({
            "doctor_id": doc_id,
            "nombre": data["nombre"],
            "total": round(data["total"], 2),
            "tratamientos": trats
        })
    ingresos_por_doctor.sort(key=lambda x: x["total"], reverse=True)

    # 2. Total payments actually collected (forma_pago != 'AB')
    q_pagos = (
        select(Pago)
        .join(Atencion)
        .where(Atencion.sucursal_id == user.sucursal_id)
        .where(Pago.forma_pago != 'AB')
    )
    if start_dt:
        q_pagos = q_pagos.where(Pago.fecha >= start_dt)
    if end_dt:
        q_pagos = q_pagos.where(Pago.fecha < end_dt)
    pagos = session.exec(q_pagos).all()
    total_cobrado = round(sum(float(p.monto) for p in pagos), 2)

    # 3. Gastos por categoría
    q_gastos = select(Gasto).where(Gasto.sucursal_id == user.sucursal_id)
    if start_dt:
        q_gastos = q_gastos.where(Gasto.fecha >= start_dt)
    if end_dt:
        q_gastos = q_gastos.where(Gasto.fecha < end_dt)
    gastos = session.exec(q_gastos).all()

    cat_map = {}
    for g in gastos:
        cat = g.categoria or "GENERAL"
        cat_map[cat] = cat_map.get(cat, 0.0) + float(g.monto)

    gastos_por_categoria = [{"categoria": k, "total": round(v, 2)} for k, v in cat_map.items()]
    gastos_por_categoria.sort(key=lambda x: x["total"], reverse=True)
    total_gastos = round(sum(g["total"] for g in gastos_por_categoria), 2)

    return {
        "ingresos": {
            "total_produccion": total_produccion,
            "total_cobrado": total_cobrado,
            "por_tratamiento": ingresos_por_tratamiento,
            "por_doctor": ingresos_por_doctor
        },
        "gastos": {
            "total": total_gastos,
            "sueldo_fijo": round(cat_map.get("SUELDO FIJO", 0.0), 2),
            "comisiones": round(cat_map.get("COMISIONES", 0.0), 2),
            # NÓMINA kept for records entered manually or migrated from older data
            "nomina_legacy": round(cat_map.get("NÓMINA", 0.0), 2),
            "insumos": round(cat_map.get("INSUMOS", 0.0), 2),
            "servicios_basicos": round(cat_map.get("SERVICIOS BÁSICOS", 0.0), 2),
            "por_categoria": gastos_por_categoria
        },
        "balance": round(total_cobrado - total_gastos, 2)
    }

@app.post("/api/pacientes/{paciente_id}/tratamientos")
def iniciar_tratamiento(paciente_id: int, tratamiento_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    # Check if already active
    existing = session.exec(select(TratamientoEnCurso).where(
        TratamientoEnCurso.paciente_id == paciente_id,
        TratamientoEnCurso.tratamiento_id == tratamiento_id,
        TratamientoEnCurso.activo == True
    )).first()
    
    if existing:
        raise HTTPException(status_code=400, detail="El paciente ya tiene este tratamiento activo")
        
    nuevo = TratamientoEnCurso(
        paciente_id=paciente_id,
        tratamiento_id=tratamiento_id,
        fecha_inicio=datetime.now(),
        activo=True
    )
    session.add(nuevo)
    session.commit()
    return {"message": "Tratamiento iniciado"}

@app.put("/api/pacientes/{paciente_id}/tratamientos/{tratamiento_id}/finalizar")
def finalizar_tratamiento(paciente_id: int, tratamiento_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    registro = session.exec(select(TratamientoEnCurso).where(
        TratamientoEnCurso.paciente_id == paciente_id,
        TratamientoEnCurso.tratamiento_id == tratamiento_id,
        TratamientoEnCurso.activo == True
    )).first()
    
    if not registro:
        raise HTTPException(status_code=404, detail="No hay tratamiento activo para finalizar")
        
    registro.activo = False
    registro.fecha_fin = datetime.now()
    session.add(registro)
    session.commit()
    return {"message": "Tratamiento finalizado"}

@app.get("/api/pacientes/{paciente_id}/tratamientos-activos")
def get_tratamientos_activos(paciente_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    registros = session.exec(select(TratamientoEnCurso).where(
        TratamientoEnCurso.paciente_id == paciente_id,
        TratamientoEnCurso.activo == True
    )).all()
    
    data = []
    for r in registros:
        data.append({
            "id": r.id,
            "tratamiento_id": r.tratamiento_id,
            "nombre": r.tratamiento.nombre if r.tratamiento else "Desconocido",
            "fecha_inicio": r.fecha_inicio
        })
    return data

class RecargaSchema(BaseModel):
    monto: float
    metodo_pago: str = "EFECTIVO" # EFECTIVO, TRANSFERENCIA, TARJETA

@app.post("/api/pacientes/{paciente_id}/recargar")
def recargar_billetera(paciente_id: int, data: RecargaSchema, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    paciente = session.get(Paciente, paciente_id)
    if not paciente:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")
    
    if data.monto <= 0:
        raise HTTPException(status_code=400, detail="El monto debe ser positivo")
        
    paciente.saldo_favor += Decimal(data.monto)
    
    # Save the recharge in the specific HistorialAbono table
    historial = HistorialAbono(
        paciente_id=paciente.id,
        usuario_id=user.id if user else None,
        monto=Decimal(data.monto),
        metodo_pago=data.metodo_pago,
        fecha=datetime.now()
    )
    session.add(historial)

    session.add(paciente)
    session.commit()
    return {"message": "Recarga exitosa", "nuevo_saldo": paciente.saldo_favor}

@app.get("/recepcion/imprimir/{atencion_id}", response_class=HTMLResponse)
def view_imprimir_atencion(atencion_id: int):
    with open("static/imprimir.html", "r", encoding="utf-8") as f:
        return f.read()

@app.get("/recepcion/editar/{atencion_id}", response_class=HTMLResponse)
def view_editar_atencion(atencion_id: int):
    with open("static/editar.html", "r", encoding="utf-8") as f:
        content = f.read()
    return HTMLResponse(content=content, headers={"Cache-Control": "no-cache, no-store, must-revalidate"})

@app.get("/recepcion", response_class=HTMLResponse)
def view_recepcion():
    with open("static/recepcion.html", "r", encoding="utf-8") as f:
        content = f.read()
    return HTMLResponse(content=content, headers={"Cache-Control": "no-cache, no-store, must-revalidate"})

@app.get("/")
def read_root():
    return {"message": "Sistema Clínica HD - FastAPI Backend Operativo", "docs": "/docs", "frontend": "/recepcion"}



RECEPCION_EDIT_HOURS = 48

def check_recepcion_time_limit(atencion: Atencion, user: User):
    """For recepcion users, blocks edits/deletes on atenciones older than RECEPCION_EDIT_HOURS."""
    if user.role == "recepcion":
        elapsed = datetime.now() - atencion.fecha
        if elapsed.total_seconds() > RECEPCION_EDIT_HOURS * 3600:
            raise HTTPException(
                status_code=403,
                detail=f"No puedes modificar esta atención. Han pasado más de {RECEPCION_EDIT_HOURS} horas desde su creación."
            )

# --- AUDITORIA HELPERS ---
def registrar_log(atencion_id: int, accion: str, descripcion: str, session: Session, user: User = None):
    log = AuditoriaAtencion(
        atencion_id=atencion_id,
        usuario_id=user.id if user else None,
        accion=accion,
        descripcion=descripcion,
        fecha=datetime.now()
    )
    session.add(log)

@app.get("/api/atenciones/{atencion_id}/historial")
def get_atencion_historial(atencion_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    logs = session.exec(
        select(AuditoriaAtencion)
        .where(AuditoriaAtencion.atencion_id == atencion_id)
        .order_by(AuditoriaAtencion.fecha.desc())
    ).all()
    
    return [
        {
            "id": l.id,
            "fecha": l.fecha,
            "accion": l.accion,
            "descripcion": l.descripcion,
            "usuario": l.usuario.username if l.usuario else "Sistema"
        } for l in logs
    ]

@app.get("/api/reportes/diario-sumario")
def get_daily_summary(session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    """
    Returns total of direct wallet recharges for today.
    """
    today = datetime.now().date()
    today_start = datetime.combine(today, datetime.min.time())
    
    # Sum recharges from AuditoriaAtencion
    logs = session.exec(
        select(AuditoriaAtencion)
        .where(AuditoriaAtencion.accion == "RECARGA_BILLETERA")
        .where(AuditoriaAtencion.fecha >= today_start)
    ).all()
    
    total_directo = 0
    for l in logs:
        try:
            # Extract amount from description "Recarga directa de $X to..."
            import re
            match = re.search(r"\$(\d+(\.\d+)?)", l.descripcion)
            if match:
                total_directo += float(match.group(1))
        except:
            pass
            
    return {"recargas_directas": total_directo}


@app.get("/api/reportes/cuadre-diario")
def get_cuadre_diario(
    fecha: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user)
):
    if not user.sucursal_id:
        raise HTTPException(status_code=400, detail="Usuario sin sucursal asignada")

    today = datetime.now().date()

    if start_date:
        date_from = datetime.combine(datetime.strptime(start_date, "%Y-%m-%d").date(), datetime.min.time())
    elif fecha:
        date_from = datetime.combine(datetime.strptime(fecha, "%Y-%m-%d").date(), datetime.min.time())
    else:
        date_from = datetime.combine(today, datetime.min.time())

    if end_date:
        date_to = datetime.combine(datetime.strptime(end_date, "%Y-%m-%d").date(), datetime.max.time())
    elif fecha:
        date_to = datetime.combine(datetime.strptime(fecha, "%Y-%m-%d").date(), datetime.max.time())
    else:
        date_to = datetime.combine(today, datetime.max.time())

    atenciones = session.exec(
        select(Atencion)
        .where(Atencion.sucursal_id == user.sucursal_id)
        .where(Atencion.fecha >= date_from)
        .where(Atencion.fecha <= date_to)
        .options(
            selectinload(Atencion.paciente),
            selectinload(Atencion.pagos),
            selectinload(Atencion.detalles)
        )
        .order_by(Atencion.fecha.asc())
    ).all()

    filas = []
    totales = {"efectivo": 0.0, "transferencia": 0.0, "tarjeta": 0.0, "abono_usado": 0.0, "total_tratamiento": 0.0}

    for a in atenciones:
        ef  = sum(float(p.monto) for p in a.pagos if p.forma_pago == "EF")
        tr  = sum(float(p.monto) for p in a.pagos if p.forma_pago == "TR")
        tc  = sum(float(p.monto) for p in a.pagos if p.forma_pago == "TC")
        ab  = sum(float(p.monto) for p in a.pagos if p.forma_pago == "AB")
        total_cobrado = ef + tr + tc + ab
        total_trat = sum(float(d.total_calculado) for d in a.detalles)
        saldo_pend = max(0.0, total_trat - total_cobrado)

        if total_trat == 0 and total_cobrado == 0:
            continue

        filas.append({
            "atencion_id": a.id,
            "fecha": a.fecha.strftime("%Y-%m-%d"),
            "paciente": f"{a.paciente.nombres} {a.paciente.apellidos}".strip() if a.paciente else "N/A",
            "historia_clinica": a.paciente.historia_clinica if a.paciente else "",
            "total_tratamiento": round(total_trat, 2),
            "efectivo":      round(ef, 2),
            "transferencia": round(tr, 2),
            "tarjeta":       round(tc, 2),
            "abono_usado":   round(ab, 2),
            "total_cobrado": round(total_cobrado, 2),
            "saldo_pendiente": round(saldo_pend, 2),
        })

        totales["efectivo"]        += ef
        totales["transferencia"]   += tr
        totales["tarjeta"]         += tc
        totales["abono_usado"]     += ab
        totales["total_tratamiento"] += total_trat

    totales["total_fisico"]  = totales["efectivo"] + totales["transferencia"] + totales["tarjeta"]
    totales["total_cobrado"] = totales["total_fisico"] + totales["abono_usado"]
    totales = {k: round(v, 2) for k, v in totales.items()}

    # Abonos generados: nuevo dinero que entró a billeteras en este período
    abonos_gen = session.exec(
        select(HistorialAbono)
        .join(Paciente, HistorialAbono.paciente_id == Paciente.id)
        .where(Paciente.sucursal_id == user.sucursal_id)
        .where(HistorialAbono.fecha >= date_from)
        .where(HistorialAbono.fecha <= date_to)
        .options(selectinload(HistorialAbono.paciente))
        .order_by(HistorialAbono.fecha.asc())
    ).all()

    abonos_lista = []
    abonos_normalizados = {"efectivo": 0.0, "transferencia": 0.0, "tarjeta": 0.0}
    for h in abonos_gen:
        metodo_raw = (h.metodo_pago or "").lower()
        monto = float(h.monto)
        if any(x in metodo_raw for x in ["efectivo", "ef", "cash"]):
            abonos_normalizados["efectivo"] += monto
        elif any(x in metodo_raw for x in ["transfer", "tr"]):
            abonos_normalizados["transferencia"] += monto
        elif any(x in metodo_raw for x in ["tarjeta", "tc", "card", "credito", "debito"]):
            abonos_normalizados["tarjeta"] += monto
        else:
            abonos_normalizados["efectivo"] += monto  # default desconocido → efectivo
        abonos_lista.append({
            "fecha":    h.fecha.strftime("%Y-%m-%d %H:%M"),
            "paciente": f"{h.paciente.nombres} {h.paciente.apellidos}".strip() if h.paciente else "N/A",
            "metodo":   h.metodo_pago or "Desconocido",
            "monto":    round(monto, 2),
        })

    abonos_normalizados = {k: round(v, 2) for k, v in abonos_normalizados.items()}

    # Tabla "Total Real de Caja": pago_tratamiento + abono_realizado por método
    real_caja = {
        "efectivo":      {"tratamiento": totales["efectivo"],      "abono": abonos_normalizados["efectivo"],      "total": round(totales["efectivo"]      + abonos_normalizados["efectivo"],      2)},
        "transferencia": {"tratamiento": totales["transferencia"],  "abono": abonos_normalizados["transferencia"], "total": round(totales["transferencia"]  + abonos_normalizados["transferencia"], 2)},
        "tarjeta":       {"tratamiento": totales["tarjeta"],        "abono": abonos_normalizados["tarjeta"],       "total": round(totales["tarjeta"]        + abonos_normalizados["tarjeta"],       2)},
    }
    real_caja["gran_total"] = {
        "tratamiento": round(totales["efectivo"] + totales["transferencia"] + totales["tarjeta"], 2),
        "abono":       round(sum(abonos_normalizados.values()), 2),
        "total":       round(totales["efectivo"] + totales["transferencia"] + totales["tarjeta"] + sum(abonos_normalizados.values()), 2),
    }

    return {
        "filas": filas,
        "totales": totales,
        "abonos_generados": abonos_lista,
        "abonos_normalizados": abonos_normalizados,
        "total_abonos_generados": round(sum(float(h.monto) for h in abonos_gen), 2),
        "real_caja": real_caja,
    }


# --- GESTIÓN DE GASTOS (Módulo 30) ---

@app.post("/api/gastos")
def create_gasto(
    descripcion: str = Form(...),
    monto: Decimal = Form(...),
    metodo_pago: str = Form(...), # EFECTIVO, TRANSFERENCIA
    categoria: str = Form("GENERAL"),
    responsable: Optional[str] = Form(None),
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user)
):
    if not user.sucursal_id:
        raise HTTPException(status_code=400, detail="El usuario no tiene una sucursal asignada")
    
    gasto = Gasto(
        descripcion=descripcion,
        monto=monto,
        metodo_pago=metodo_pago,
        categoria=categoria,
        responsable=responsable,
        sucursal_id=user.sucursal_id,
        usuario_id=user.id
    )
    session.add(gasto)
    session.commit()
    session.refresh(gasto)
    return gasto

@app.get("/api/gastos")
def list_gastos(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    search: Optional[str] = None,
    categoria: Optional[str] = None,
    metodo_pago: Optional[str] = None,
    usuario_id: Optional[int] = None,
    page: int = 1,
    size: int = 50,
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user)
):
    skip = (page - 1) * size
    if not user.sucursal_id:
        return []
    
    # Eager load 'usuario' to show who registered the expense
    query = (
        select(Gasto)
        .where(Gasto.sucursal_id == user.sucursal_id)
        .options(selectinload(Gasto.usuario))
        .order_by(Gasto.fecha.desc())
    )
    
    if start_date:
        query = query.where(Gasto.fecha >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        query = query.where(Gasto.fecha < end_dt)
    
    if search:
        query = query.where(Gasto.descripcion.ilike(f"%{search}%"))
        
    if categoria:
        query = query.where(Gasto.categoria == categoria)
        
    if metodo_pago:
        query = query.where(Gasto.metodo_pago == metodo_pago)
        
    if usuario_id:
        query = query.where(Gasto.usuario_id == usuario_id)
        
    return session.exec(query.offset(skip).limit(size)).all()

@app.get("/api/gastos/balances")
def get_gastos_balances(session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    """
    Calculates Strict Financial Balances:
    Caja Fija = Pagos(EF) + Recargas(EFECTIVO) - Gastos(EFECTIVO)
    Banco = Pagos(TR) + Recargas(TRANSFERENCIA) - Gastos(TRANSFERENCIA)
    Tarjeta = Pagos(TC) + Recargas(TARJETA)
    Billetera Aplicada = Pagos(AB)
    """
    if not user.sucursal_id:
         return {"efectivo": 0, "transferencia": 0, "tarjeta": 0, "abono_aplicado": 0}
    
    # --- 1. INCOME FROM PAYMENTS (Atenciones) ---
    ingresos_query = session.exec(
        select(Pago.forma_pago, func.sum(Pago.monto))
        .join(Atencion)
        .where(Atencion.sucursal_id == user.sucursal_id)
        .group_by(Pago.forma_pago)
    ).all()
    
    ingresos_crudos = {row[0]: float(row[1]) for row in ingresos_query}
    
    # Map to standard concepts
    ingresos = {
        "EFECTIVO": ingresos_crudos.get("EF", 0),
        "TRANSFERENCIA": ingresos_crudos.get("TR", 0),
        "TARJETA": ingresos_crudos.get("TC", 0),
        "ABONO_APLICADO": ingresos_crudos.get("AB", 0) # Internal movement
    }
    
    # --- 2. INCOME FROM WALLET RECHARGES (Recargas Directas) ---
    recargas_query = session.exec(
        select(HistorialAbono.metodo_pago, func.sum(HistorialAbono.monto))
        .join(User, HistorialAbono.usuario_id == User.id)
        .where(User.sucursal_id == user.sucursal_id)
        .group_by(HistorialAbono.metodo_pago)
    ).all()
    
    recargas = {row[0]: float(row[1]) for row in recargas_query}

    # Add Recharges to gross income
    ingresos["EFECTIVO"] += recargas.get("EFECTIVO", 0)
    ingresos["TRANSFERENCIA"] += recargas.get("TRANSFERENCIA", 0)
    ingresos["TARJETA"] += recargas.get("TARJETA", 0)

    # --- 3. EXPENSES (Gastos) ---
    egresos_query = session.exec(
        select(Gasto.metodo_pago, func.sum(Gasto.monto))
        .where(Gasto.sucursal_id == user.sucursal_id)
        .group_by(Gasto.metodo_pago)
    ).all()
    
    egresos = {row[0]: float(row[1]) for row in egresos_query}
    
    # --- 4. CALCULATE STRICT BALANCES ---
    balance_efectivo = ingresos["EFECTIVO"] - egresos.get("EFECTIVO", 0)
    balance_transferencia = ingresos["TRANSFERENCIA"] - egresos.get("TRANSFERENCIA", 0)
    balance_tarjeta = ingresos["TARJETA"] # No expenses in card currently
    
    return {
        "efectivo": balance_efectivo,
        "transferencia": balance_transferencia,
        "tarjeta": balance_tarjeta,
        "abono_aplicado": ingresos["ABONO_APLICADO"],
        "desglose_ingresos": ingresos,
        "desglose_egresos": egresos,
        "desglose_recargas": recargas
    }

# --- SISTEMA DE NÓMINA Y COMISIONES (Módulo 35) ---

def calcular_cascada_paciente(session: Session, pacientes_ids: set, sucursal_id: int = None, start_date: str = None, end_date: str = None, empleado_id: int = None, tipo_empleado: str = None):
    """
    Lógica Núcleo de la Cascada FIFO.
    Agrupa los pagos globales de cada paciente y los distribuye cronológicamente sobre sus tratamientos.
    """
    resultados = []
    start_dt = datetime.strptime(start_date, "%Y-%m-%d") if start_date else None
    end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) if end_date else None

    for pac_id in pacientes_ids:
        atenciones = session.exec(
            select(Atencion)
            .where(Atencion.paciente_id == pac_id)
            .options(
                selectinload(Atencion.pagos), 
                selectinload(Atencion.detalles).selectinload(AtencionDetalle.tratamiento)
            )
        ).all()
        
        atenciones.sort(key=lambda a: a.fecha)
        pool_pagos = sum(float(p.monto) for a in atenciones for p in a.pagos)
        
        for atencion in atenciones:
            if not atencion.validado:
                continue
                
            detalles = sorted(atencion.detalles, key=lambda d: d.id)
            suma_calculada_detalles = sum(float(d.total_calculado) for d in detalles)
            
            total_atn = float(atencion.total_atencion_valor) if getattr(atencion, 'total_atencion_valor', None) else 0.0
            factor = (total_atn / suma_calculada_detalles) if suma_calculada_detalles > 0 else 1.0
            
            for detalle in detalles:
                costo_efectivo = float(detalle.total_calculado) * factor
                
                cobrado = min(pool_pagos, costo_efectivo)
                pool_pagos -= cobrado
                
                porcentaje_detalle = (cobrado / costo_efectivo) if costo_efectivo > 0 else 1.0
                
                comision_teorica = float(detalle.comision_valor)
                comision_real_total = comision_teorica * porcentaje_detalle
                comision_ya_pagada = float(detalle.comision_pagada_monto) if getattr(detalle, 'comision_pagada_monto', None) else 0.0
                comision_pendiente = comision_real_total - comision_ya_pagada
                
                if comision_pendiente > 0.01:
                    if sucursal_id and atencion.sucursal_id != sucursal_id:
                        continue
                    if start_dt and atencion.fecha < start_dt:
                        continue
                    if end_dt and atencion.fecha >= end_dt:
                        continue
                        
                    if empleado_id and tipo_empleado:
                        if tipo_empleado == 'Doctor' and detalle.doctor_id != empleado_id:
                            continue
                        if tipo_empleado == 'Secretaria' and detalle.vendedor_id != empleado_id:
                            continue
                            
                    resultados.append({
                        "detalle": detalle,
                        "comision_pendiente": comision_pendiente,
                        "comision_real_total": comision_real_total,
                        "porcentaje_pago_paciente": porcentaje_detalle,
                        "atencion": atencion
                    })
    return resultados

def get_cascada_pendientes_global(session: Session, sucursal_id: int = None, start_date: str = None, end_date: str = None, empleado_id: int = None, tipo_empleado: str = None):
    query_pendientes = select(AtencionDetalle).join(Atencion).where(Atencion.validado == True).where(AtencionDetalle.comision_pagada == False)
    if sucursal_id:
        query_pendientes = query_pendientes.where(Atencion.sucursal_id == sucursal_id)
        
    pendientes_crudos = session.exec(query_pendientes).all()
    if not pendientes_crudos:
        return []
        
    pacientes_ids = {p.atencion.paciente_id for p in pendientes_crudos if p.atencion}
    return calcular_cascada_paciente(session, pacientes_ids, sucursal_id, start_date, end_date, empleado_id, tipo_empleado)

@app.get("/api/nomina/pendientes")
def get_nomina_pendientes(start_date: str = None, end_date: str = None, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    """
    Returns pending commissions iteratively via FIFO across all patient's attentions.
    """
    if not user.sucursal_id:
        raise HTTPException(status_code=400, detail="El usuario no tiene una sucursal asignada")
    
    doctores_pendientes = {}
    doc_query = select(Doctor).where(Doctor.activo == True)
    if user.sucursal_id:
        doc_query = doc_query.where((Doctor.sucursal_id == user.sucursal_id) | (Doctor.sucursal_id == None))
    all_docs = session.exec(doc_query).all()
    
    for doc in all_docs:
        doctores_pendientes[doc.id] = {
            "id": doc.id,
            "nombre": f"{doc.nombres} {doc.apellidos}",
            "tipo": "Doctor",
            "comisiones_acumuladas": 0.0,
            "detalles": []
        }
    
    vendedores_pendientes = {}
    
    cascada_results = get_cascada_pendientes_global(session, user.sucursal_id, start_date, end_date)
    
    for res in cascada_results:
        d = res["detalle"]
        com_pen = res["comision_pendiente"]
        
        # Doc logic
        if d.doctor_id and d.doctor_id in doctores_pendientes:
            doctores_pendientes[d.doctor_id]["comisiones_acumuladas"] += com_pen
            doctores_pendientes[d.doctor_id]["detalles"].append({
                "detalle_id": d.id,
                "tratamiento": d.tratamiento.nombre if d.tratamiento else "Desconocido",
                "valor_tratamiento": float(d.total_calculado),
                "porcentaje": float(d.porcentaje_comision),
                "comision": round(com_pen, 2),
                "fecha": res["atencion"].fecha.isoformat(),
                "porcentaje_pago_real": res["porcentaje_pago_paciente"]
            })
            
        # Sec logic
        if d.vendedor_id:
            if d.vendedor_id not in vendedores_pendientes:
                vendedor_user = session.get(User, d.vendedor_id)
                vendedores_pendientes[d.vendedor_id] = {
                    "id": d.vendedor_id,
                    "nombre": vendedor_user.username if vendedor_user else "Usuario",
                    "tipo": "Secretaria/Ventas",
                    "comisiones_acumuladas": 0.0,
                    "detalles": []
                }
            vendedores_pendientes[d.vendedor_id]["comisiones_acumuladas"] += com_pen
            vendedores_pendientes[d.vendedor_id]["detalles"].append({
                "detalle_id": d.id,
                "tratamiento": d.tratamiento.nombre if d.tratamiento else "Venta",
                "valor_tratamiento": float(d.total_calculado),
                "porcentaje": float(d.porcentaje_comision),
                "comision": round(com_pen, 2),
                "fecha": res["atencion"].fecha.isoformat(),
                "porcentaje_pago_real": res["porcentaje_pago_paciente"]
            })
            
    return {
        "doctores": list(doctores_pendientes.values()),
        "vendedores": list(vendedores_pendientes.values())
    }

class PagoNominaSchema(BaseModel):
    empleado_id: int
    tipo_empleado: str # 'Doctor' or 'Secretaria'
    sueldo_base: float = 0.0
    comisiones: float = 0.0
    efectivo: float = 0.0
    transferencia: float = 0.0
    tarjeta: float = 0.0
    responsable: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    
@app.post("/api/nomina/pagar")
def pagar_nomina(data: PagoNominaSchema, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    if not user.sucursal_id:
        raise HTTPException(status_code=400, detail="El usuario no tiene una sucursal asignada")
        
    efectivo_val = Decimal(str(data.efectivo))
    transferencia_val = Decimal(str(data.transferencia))
    tarjeta_val = Decimal(str(data.tarjeta))
    total_entregado = efectivo_val + transferencia_val + tarjeta_val
    total_a_pagar = Decimal(str(data.sueldo_base)) + Decimal(str(data.comisiones))
    
    if total_a_pagar <= 0:
        raise HTTPException(status_code=400, detail="El total a pagar debe ser mayor a 0")
        
    if abs(total_entregado - total_a_pagar) > Decimal("0.02"):
        raise HTTPException(status_code=400, detail=f"La suma de métodos de pago debe ser igual al Total a Pagar (${total_a_pagar})")
        
    sucursal = session.get(Sucursal, user.sucursal_id)
    if not sucursal:
        raise HTTPException(status_code=400, detail="Sucursal no encontrada")
        
    balances = get_gastos_balances(session=session, user=user)
        
    if Decimal(str(balances["efectivo"])) < efectivo_val:
        raise HTTPException(status_code=400, detail=f"Fondos insuficientes en Efectivo. Disponible: ${balances['efectivo']}")
        
    if Decimal(str(balances["transferencia"])) < transferencia_val:
        raise HTTPException(status_code=400, detail=f"Fondos insuficientes en Bancos. Disponible: ${balances['transferencia']}")
        
    if Decimal(str(balances["tarjeta"])) < tarjeta_val:
        raise HTTPException(status_code=400, detail=f"Fondos insuficientes en Tarjeta. Disponible: ${balances['tarjeta']}")
        
    now = datetime.now()
    empleado_nombre = "Empleado"
    
    if data.tipo_empleado == 'Doctor':
        doctor = session.get(Doctor, data.empleado_id)
        if doctor: empleado_nombre = f"Dr(a). {doctor.nombres} {doctor.apellidos}"
    elif data.tipo_empleado == 'Secretaria':
        secretaria = session.get(User, data.empleado_id)
        if secretaria: empleado_nombre = f"Sec. {secretaria.username}"

    # Extract pending details with proper FIFO mapping matching exactly what was visualized
    cascada_results = get_cascada_pendientes_global(
        session=session, 
        sucursal_id=user.sucursal_id, 
        start_date=data.start_date, 
        end_date=data.end_date,
        empleado_id=data.empleado_id,
        tipo_empleado=data.tipo_empleado
    )
    
    monto_a_repartir = float(data.comisiones)
    
    for res in cascada_results:
        d = res["detalle"]
        com_pen = res["comision_pendiente"]
        pct_pago = res["porcentaje_pago_paciente"]
        
        if monto_a_repartir <= 0.001:
            break
            
        pago_para_este_detalle = min(monto_a_repartir, com_pen)
        monto_a_repartir -= pago_para_este_detalle
        
        d.comision_pagada_monto = float(getattr(d, 'comision_pagada_monto', 0.0) or 0.0) + pago_para_este_detalle
        d.fecha_pago_comision = now
        
        # Treatment is fully paid by patient AND the clinic fully paid the doctor for the theoretical amount
        if pct_pago >= 0.99 and abs(float(d.comision_pagada_monto) - float(d.comision_valor)) < 0.02:
            d.comision_pagada = True
            
        session.add(d)
    
    # Create Gasto records split by concept (SUELDO FIJO / COMISIONES) and payment method.
    # Each payment method amount is split proportionally between both concepts.
    total_a_pagar_f = float(total_a_pagar)
    pct_sueldo = data.sueldo_base / total_a_pagar_f if total_a_pagar_f > 0 else 0.0
    pct_com    = data.comisiones  / total_a_pagar_f if total_a_pagar_f > 0 else 0.0

    desc_sueldo    = f"Sueldo Fijo: {empleado_nombre}"
    desc_comision  = f"Comisiones: {empleado_nombre}"

    def _registrar_split(metodo: str, monto_metodo: Decimal):
        if monto_metodo <= 0:
            return
        monto_f = float(monto_metodo)
        if data.sueldo_base > 0:
            monto_s = round(monto_f * pct_sueldo, 2)
            if monto_s > 0:
                session.add(Gasto(
                    fecha=now, descripcion=desc_sueldo,
                    monto=Decimal(str(monto_s)), metodo_pago=metodo,
                    categoria="SUELDO FIJO",
                    responsable=data.responsable or user.username,
                    sucursal_id=user.sucursal_id, usuario_id=user.id
                ))
        if data.comisiones > 0:
            monto_c = round(monto_f * pct_com, 2)
            if monto_c > 0:
                session.add(Gasto(
                    fecha=now, descripcion=desc_comision,
                    monto=Decimal(str(monto_c)), metodo_pago=metodo,
                    categoria="COMISIONES",
                    responsable=data.responsable or user.username,
                    sucursal_id=user.sucursal_id, usuario_id=user.id
                ))

    _registrar_split("EFECTIVO",      efectivo_val)
    _registrar_split("TRANSFERENCIA", transferencia_val)
    _registrar_split("TARJETA",       tarjeta_val)
        
    session.commit()
    
    return {"message": f"Nómina pagada exitosamente a {empleado_nombre}", "total_pagado": float(total_entregado)}

class DetalleSeleccionItem(BaseModel):
    detalle_id: int
    comision: float

class PagoNominaSeleccionSchema(BaseModel):
    empleado_id: int
    tipo_empleado: str  # 'Doctor' or 'Secretaria'
    detalles: List[DetalleSeleccionItem]
    efectivo: float = 0.0
    transferencia: float = 0.0
    tarjeta: float = 0.0
    responsable: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None

@app.post("/api/nomina/pagar-seleccion")
def pagar_nomina_seleccion(data: PagoNominaSeleccionSchema, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    if not user.sucursal_id:
        raise HTTPException(status_code=400, detail="El usuario no tiene una sucursal asignada")
    if not data.detalles:
        raise HTTPException(status_code=400, detail="No se seleccionaron tratamientos")

    efectivo_val = Decimal(str(data.efectivo))
    transferencia_val = Decimal(str(data.transferencia))
    tarjeta_val = Decimal(str(data.tarjeta))
    total_entregado = efectivo_val + transferencia_val + tarjeta_val
    total_seleccion = Decimal(str(round(sum(item.comision for item in data.detalles), 2)))

    if total_seleccion <= 0:
        raise HTTPException(status_code=400, detail="El total a pagar debe ser mayor a 0")
    if abs(total_entregado - total_seleccion) > Decimal("0.05"):
        raise HTTPException(status_code=400, detail=f"La suma de métodos (${total_entregado}) no coincide con el total seleccionado (${total_seleccion})")

    balances = get_gastos_balances(session=session, user=user)
    if Decimal(str(balances["efectivo"])) < efectivo_val:
        raise HTTPException(status_code=400, detail=f"Fondos insuficientes en Efectivo. Disponible: ${balances['efectivo']}")
    if Decimal(str(balances["transferencia"])) < transferencia_val:
        raise HTTPException(status_code=400, detail=f"Fondos insuficientes en Bancos. Disponible: ${balances['transferencia']}")
    if Decimal(str(balances["tarjeta"])) < tarjeta_val:
        raise HTTPException(status_code=400, detail=f"Fondos insuficientes en Tarjeta. Disponible: ${balances['tarjeta']}")

    now = datetime.now()
    empleado_nombre = "Empleado"
    if data.tipo_empleado == 'Doctor':
        doctor = session.get(Doctor, data.empleado_id)
        if doctor: empleado_nombre = f"Dr(a). {doctor.nombres} {doctor.apellidos}"
    elif data.tipo_empleado == 'Secretaria':
        sec = session.get(User, data.empleado_id)
        if sec: empleado_nombre = f"Sec. {sec.username}"

    # Fetch cascada to get current state (porcentaje_pago_paciente, comision_pendiente per detalle)
    cascada_results = get_cascada_pendientes_global(
        session=session,
        sucursal_id=user.sucursal_id,
        empleado_id=data.empleado_id,
        tipo_empleado=data.tipo_empleado
    )
    cascada_by_id = {res["detalle"].id: res for res in cascada_results}

    for item in data.detalles:
        if item.detalle_id not in cascada_by_id:
            raise HTTPException(status_code=400, detail=f"El tratamiento ID {item.detalle_id} no está pendiente o no existe")
        res = cascada_by_id[item.detalle_id]
        d = res["detalle"]
        com_pen = res["comision_pendiente"]
        pct_pago = res["porcentaje_pago_paciente"]

        pago = min(item.comision, com_pen)
        d.comision_pagada_monto = float(getattr(d, 'comision_pagada_monto', 0.0) or 0.0) + pago
        d.fecha_pago_comision = now
        if pct_pago >= 0.99 and abs(float(d.comision_pagada_monto) - float(d.comision_valor)) < 0.02:
            d.comision_pagada = True
        session.add(d)

    desc_comision = f"Comisiones: {empleado_nombre}"
    def _reg(metodo: str, monto: Decimal):
        if monto > 0:
            session.add(Gasto(
                fecha=now, descripcion=desc_comision,
                monto=monto, metodo_pago=metodo,
                categoria="COMISIONES",
                responsable=data.responsable or user.username,
                sucursal_id=user.sucursal_id, usuario_id=user.id
            ))
    _reg("EFECTIVO", efectivo_val)
    _reg("TRANSFERENCIA", transferencia_val)
    _reg("TARJETA", tarjeta_val)

    session.commit()
    return {"message": f"{len(data.detalles)} tratamiento(s) pagado(s) a {empleado_nombre}", "total_pagado": float(total_entregado)}

class RetiroSociosSchema(BaseModel):
    monto: float
    metodo_pago: str # 'EFECTIVO' or 'TRANSFERENCIA'
    descripcion: str
    responsable: Optional[str] = None

@app.post("/api/nomina/retiro-socios")
def retiro_socios(data: RetiroSociosSchema, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    """ Endpoint for admins to withdraw net profits """
    if user.role != "admin":
         raise HTTPException(status_code=403, detail="Solo los administradores pueden registrar retiros de socios")
         
    if not user.sucursal_id:
         raise HTTPException(status_code=400, detail="Sin sucursal asignada")
         
    if data.monto <= 0:
         raise HTTPException(status_code=400, detail="El monto debe ser positivo")
         
    sucursal = session.get(Sucursal, user.sucursal_id)
    if not sucursal:
        raise HTTPException(status_code=400, detail="Sucursal no encontrada")
        
    balances = get_gastos_balances(session=session, user=user)
    monto_dec = Decimal(str(data.monto))
    
    if data.metodo_pago == "EFECTIVO" and Decimal(str(balances["efectivo"])) < monto_dec:
        raise HTTPException(status_code=400, detail=f"Fondos insuficientes en Efectivo. Disponible: ${balances['efectivo']}")
        
    if data.metodo_pago == "TRANSFERENCIA" and Decimal(str(balances["transferencia"])) < monto_dec:
        raise HTTPException(status_code=400, detail=f"Fondos insuficientes en Bancos. Disponible: ${balances['transferencia']}")
        
    if data.metodo_pago == "TARJETA" and Decimal(str(balances["tarjeta"])) < monto_dec:
        raise HTTPException(status_code=400, detail=f"Fondos insuficientes en Tarjeta. Disponible: ${balances['tarjeta']}")
         
    gasto = Gasto(
        fecha=datetime.now(),
        descripcion=f"Retiro de Socios / Utilidades: {data.descripcion}",
        monto=Decimal(data.monto),
        metodo_pago=data.metodo_pago,
        categoria="RETIRO SOCIOS",
        responsable=data.responsable or user.username,
        sucursal_id=user.sucursal_id,
        usuario_id=user.id
    )
    session.add(gasto)
    session.commit()
    return {"message": "Retiro registrado exitosamente como Gasto"}
    
class TransferenciaInternaSchema(BaseModel):
    desde: str
    hacia: str
    monto: float
    descripcion: str
    responsable: str

@app.post("/api/gastos/transferencia")
def internal_transfer(data: TransferenciaInternaSchema, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    """ Moves funds between methods by creating two mirrored expense entries """
    if not user.sucursal_id:
        raise HTTPException(status_code=400, detail="Sin sucursal asignada")
    
    if data.desde == data.hacia:
        raise HTTPException(status_code=400, detail="Origen y destino no pueden ser iguales")
    
    if data.monto <= 0:
        raise HTTPException(status_code=400, detail="El monto debe ser positivo")

    # 1. Negative entry (Income for the destination)
    # Since balances = Income - Expenses, a negative expense is an Income.
    entry_to = Gasto(
        fecha=datetime.now(),
        descripcion=f"RECEPCIÓN TRANSFERENCIA: {data.descripcion}",
        monto=Decimal(-data.monto),
        metodo_pago=data.hacia,
        categoria="TRANSFERENCIA INTERNA",
        responsable=data.responsable,
        sucursal_id=user.sucursal_id,
        usuario_id=user.id
    )
    
    # 2. Positive entry (Expense for the origin)
    entry_from = Gasto(
        fecha=datetime.now(),
        descripcion=f"SALIDA TRANSFERENCIA: {data.descripcion}",
        monto=Decimal(data.monto),
        metodo_pago=data.desde,
        categoria="TRANSFERENCIA INTERNA",
        responsable=data.responsable,
        sucursal_id=user.sucursal_id,
        usuario_id=user.id
    )

    session.add(entry_to)
    session.add(entry_from)
    session.commit()
    
    return {"message": "Transferencia realizada con éxito"}

# --- FIN SISTEMA DE NÓMINA ---

@app.get("/api/atenciones/historial/global")
def get_global_historial(page: int = 1, size: int = 50, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    try:
        skip = (page - 1) * size
        # Join with Atencion and Paciente to provide more context in the global list
        logs = session.exec(
            select(AuditoriaAtencion)
            .order_by(AuditoriaAtencion.fecha.desc())
            .offset(skip).limit(size)
        ).all()
        
        return [
            {
                "id": l.id,
                "fecha": l.fecha,
                "atencion_id": l.atencion_id,
                "paciente_nombre": f"{l.atencion.paciente.nombres} {l.atencion.paciente.apellidos}" if l.atencion and l.atencion.paciente else "N/A",
                "accion": l.accion,
                "descripcion": l.descripcion,
                "usuario": l.usuario.username if l.usuario else "Sistema"
            } for l in logs
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/recepcion/historial", response_class=HTMLResponse)
def view_historial():
    with open("static/historial.html", "r", encoding="utf-8") as f:
        return f.read()

@app.get("/api/secret/patch-db")
def secret_patch_db(session: Session = Depends(get_session)):
    from sqlalchemy import text
    try:
        session.execute(text("ALTER TABLE paciente ADD COLUMN sucursal_id INTEGER REFERENCES sucursal (id);"))
        session.commit()
    except Exception:
        session.rollback()

    target_name = "Healthy Dental La Magdalena Sur"
    sucursal = session.exec(select(Sucursal).where(Sucursal.nombre == target_name)).first()
    if not sucursal:
        sucursal = Sucursal(nombre=target_name, direccion="Sur")
        session.add(sucursal)
        session.commit()
        session.refresh(sucursal)

    pacientes = session.exec(select(Paciente)).all()
    count = 0
    for p in pacientes:
        if not p.sucursal_id:
            p.sucursal_id = sucursal.id
        suc = session.get(Sucursal, p.sucursal_id)
        prefix = suc.nombre[:3].upper() if suc and len(suc.nombre) >= 3 else "GEN"
        expected_hc = f"HC-{prefix}-{p.id:04d}"
        if p.historia_clinica != expected_hc:
            p.historia_clinica = expected_hc
            count += 1
        session.add(p)
    if count > 0:
        session.commit()
        
    return {"status": "ok", "patched_pacientes": count, "sucursal_id": sucursal.id}


# --- TEMP: eliminar atencion NELI CHURACO (HC-EL -0576) mayo 16 ---
@app.get("/api/temp/del-atencion-neli-may16")
def temp_del_atencion_neli(clave: str, session: Session = Depends(get_session)):
    if clave != "hd2026fix":
        raise HTTPException(status_code=403, detail="Clave incorrecta")
    paciente = session.exec(select(Paciente).where(Paciente.historia_clinica == "HC-EL -0576")).first()
    if not paciente:
        return {"error": "Paciente HC-EL -0576 no encontrado"}
    fecha_start = datetime(2026, 5, 16, 0, 0, 0)
    fecha_end   = datetime(2026, 5, 16, 23, 59, 59)
    atenciones = session.exec(
        select(Atencion)
        .where(Atencion.paciente_id == paciente.id)
        .where(Atencion.fecha >= fecha_start)
        .where(Atencion.fecha <= fecha_end)
    ).all()
    if not atenciones:
        return {"mensaje": "No se encontraron atenciones de HC-EL -0576 con fecha 2026-05-16"}
    eliminadas = []
    for a in atenciones:
        for pago in a.pagos:
            session.delete(pago)
        for detalle in a.detalles:
            session.delete(detalle)
        session.flush()
        eliminadas.append({"atencion_id": a.id, "fecha": str(a.fecha)})
        session.delete(a)
    session.commit()
    return {"ok": True, "eliminadas": eliminadas}
# --- END TEMP ---

# --- END API ROUTES ---

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
